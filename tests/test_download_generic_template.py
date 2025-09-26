# NG-HEADER: Nombre de archivo: test_download_generic_template.py
# NG-HEADER: Ubicación: tests/test_download_generic_template.py
# NG-HEADER: Descripción: Pruebas de descarga de plantillas genéricas.
# NG-HEADER: Lineamientos: Ver AGENTS.md
import os
import asyncio
from io import BytesIO

from openpyxl import load_workbook
from fastapi.testclient import TestClient

os.environ["DB_URL"] = "sqlite+aiosqlite:///:memory:"
os.environ["SECRET_KEY"] = "test"
os.environ["ADMIN_PASS"] = "test"
os.environ["AUTH_ENABLED"] = "true"

from services.api import app  # noqa: E402
from db.base import Base  # noqa: E402
from db.session import engine  # noqa: E402
from services.auth import SessionData, current_session  # noqa: E402


async def _init_db() -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)


asyncio.get_event_loop().run_until_complete(_init_db())

client = TestClient(app)
app.dependency_overrides[current_session] = lambda: SessionData(None, None, "admin")


def test_download_generic_template() -> None:
    resp = client.get("/suppliers/price-list/template")
    assert resp.status_code == 200
    assert "plantilla-generica.xlsx" in resp.headers.get("content-disposition", "")
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.title == "data"
    headers = [
        "ID",
        "Agrupamiento",
        "Familia",
        "SubFamilia",
        "Producto",
        "Compra Minima",
        "Stock",
        "PrecioDeCompra",
        "PrecioDeVenta",
    ]
    assert [cell.value for cell in ws[1]] == headers


def test_download_generic_template_unauthorized() -> None:
    app.dependency_overrides[current_session] = lambda: SessionData(None, None, "guest")
    resp = client.get("/suppliers/price-list/template")
    assert resp.status_code == 403
    app.dependency_overrides[current_session] = lambda: SessionData(None, None, "admin")
