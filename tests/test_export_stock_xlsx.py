#!/usr/bin/env python
# NG-HEADER: Nombre de archivo: test_export_stock_xlsx.py
# NG-HEADER: Ubicación: tests/test_export_stock_xlsx.py
# NG-HEADER: Descripción: Pruebas del exportador XLS de stock (prioridad canónica y estilos básicos).
# NG-HEADER: Lineamientos: Ver AGENTS.md
import io
import os
import pytest
from sqlalchemy import select

# Forzar entorno de pruebas aislado (SQLite en memoria)
os.environ["DB_URL"] = "sqlite+aiosqlite:///:memory:"
os.environ["SECRET_KEY"] = "test"
os.environ["AUTH_ENABLED"] = "true"

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from services.api import app
from services.auth import current_session, require_csrf, SessionData
from db.models import Category, Product, Supplier, SupplierProduct, ProductEquivalence, CanonicalProduct, StockLedger

client = TestClient(app)
app.dependency_overrides[current_session] = lambda: SessionData(None, None, "admin")
app.dependency_overrides[require_csrf] = lambda: None


async def _seed_basic():
    from db.session import SessionLocal
    async with SessionLocal() as s:  # type: ignore
        sup = Supplier(slug="acme", name="ACME")
        category = Category(name="Cultivo", kind="category")
        subcategory = Category(name="Interior", kind="subcategory")
        s.add_all([sup, category, subcategory])
        await s.flush()
        p = Product(sku_root="SKU1", title="Producto X", category_id=category.id, subcategory_id=subcategory.id)
        s.add(p)
        await s.flush()
        sp = SupplierProduct(supplier_id=sup.id, supplier_product_id="P1", title="Prod X", internal_product_id=p.id)
        sp.current_sale_price = 123.45
        s.add(sp)
        await s.flush()
        cp = CanonicalProduct(name="Canon X", sku_custom="AAA_0001_BBB", category_id=category.id, subcategory_id=subcategory.id)
        cp.sale_price = 99.99
        s.add(cp)
        await s.flush()
        eq = ProductEquivalence(supplier_id=sup.id, supplier_product_id=sp.id, canonical_product_id=cp.id, source="test")
        s.add(eq)
        await s.commit()
        return p.id

@pytest.mark.asyncio
async def test_exporter_prefers_canonical_and_styles_header():
    await _seed_basic()
    r = client.get("/stock/export.xlsx")
    assert r.status_code == 200
    content = r.content
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # Encabezados
    headers = [c.value for c in ws[1]]
    assert headers == ["NOMBRE DE PRODUCTO", "PRECIO DE VENTA", "CATEGORIA", "SKU PROPIO"]
    # Buscar la fila del producto sembrado (Canon X)
    found = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Canon X":
            found = r
            break
    assert found is not None, "No se encontró la fila de Canon X en el XLS"
    assert float(ws.cell(row=found, column=2).value) == 99.99  # precio canónico domina
    assert ws.cell(row=found, column=3).value == "Cultivo > Interior"
    # Estilo básico del header (negrita)
    assert ws.cell(row=1, column=1).font.bold is True


@pytest.mark.asyncio
async def test_csv_and_pdf_share_stock_export_contract():
    await _seed_basic()

    csv_response = client.get("/stock/export.csv?stock=eq:0")
    assert csv_response.status_code == 200
    assert csv_response.headers["content-type"].startswith("text/csv")
    csv_text = csv_response.content.decode("utf-8-sig")
    assert "NOMBRE DE PRODUCTO,PRECIO DE VENTA,CATEGORIA,SKU PROPIO" in csv_text
    assert "Canon X,99.99,Cultivo > Interior,AAA_0001_BBB" in csv_text

    pdf_response = client.get("/stock/export.pdf?stock=eq:0")
    assert pdf_response.status_code == 200
    assert pdf_response.headers["content-type"] == "application/pdf"
    assert pdf_response.content.startswith(b"%PDF")


@pytest.mark.asyncio
async def test_manual_stock_adjustment_is_decimal_audited_and_conflict_safe():
    product_id = await _seed_basic()

    response = client.patch(
        f"/products/{product_id}/stock",
        json={"stock": 10.25, "expected_stock": 0},
    )
    assert response.status_code == 200, response.text
    assert response.json()["stock"] == 10.25

    stale = client.patch(
        f"/products/{product_id}/stock",
        json={"stock": 11, "expected_stock": 0},
    )
    assert stale.status_code == 409
    assert stale.json()["detail"]["current_stock"] == 10.25

    from db.session import SessionLocal
    async with SessionLocal() as session:
        movements = (
            await session.execute(
                select(StockLedger).where(
                    StockLedger.product_id == product_id,
                    StockLedger.source_type == "manual_adjustment",
                )
            )
        ).scalars().all()
        assert len(movements) == 1
        assert float(movements[0].delta) == 10.25
        assert float(movements[0].balance_after) == 10.25
