#!/usr/bin/env python
# NG-HEADER: Nombre de archivo: test_export_tiendanegocio_xlsx.py
# NG-HEADER: Ubicación: tests/test_export_tiendanegocio_xlsx.py
# NG-HEADER: Descripción: Prueba del export XLSX en formato TiendaNegocio.
# NG-HEADER: Lineamientos: Ver AGENTS.md
import io
import os
import pytest

# Forzar entorno de pruebas aislado (SQLite en memoria)
os.environ["DB_URL"] = "sqlite+aiosqlite:///:memory:"
os.environ["SECRET_KEY"] = "test"
os.environ["AUTH_ENABLED"] = "true"

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from services.api import app
from services.auth import current_session, require_csrf, SessionData
from db.models import Product, Supplier, SupplierProduct, ProductEquivalence, CanonicalProduct

client = TestClient(app)
app.dependency_overrides[current_session] = lambda: SessionData(None, None, "admin")
app.dependency_overrides[require_csrf] = lambda: None


async def _seed_basic():
    from db.session import SessionLocal
    async with SessionLocal() as s:  # type: ignore
        sup = Supplier(slug="acme", name="ACME")
        s.add(sup)
        await s.flush()
        p = Product(sku_root="SKU1", title="Producto X", stock=7)
        # agregar datos técnicos
        p.weight_kg = 0.5
        p.height_cm = 10
        p.width_cm = 20
        p.depth_cm = 30
        s.add(p)
        await s.flush()
        sp = SupplierProduct(supplier_id=sup.id, supplier_product_id="P1", title="Prod X", internal_product_id=p.id)
        sp.current_sale_price = 123.45
        s.add(sp)
        await s.flush()
        cp = CanonicalProduct(name="Canon X", sku_custom="AAA_0001_BBB")
        cp.sale_price = 99.99
        s.add(cp)
        await s.flush()
        eq = ProductEquivalence(supplier_id=sup.id, supplier_product_id=sp.id, canonical_product_id=cp.id, source="test")
        s.add(eq)
        await s.commit()
        return p.id


@pytest.mark.asyncio
async def test_export_tiendanegocio_headers_and_values():
    await _seed_basic()
    r = client.get("/stock/export-tiendanegocio.xlsx")
    assert r.status_code == 200
    content = r.content
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # Encabezados
    headers = [c.value for c in ws[1]]
    assert headers == [
        "SKU (OBLIGATORIO)",
        "Nombre del producto",
        "Precio",
        "Oferta",
        "Stock",
        "Visibilidad (Visible o Oculto)",
        "Descripción",
        "Peso en KG",
        "Alto en CM",
        "Ancho en CM",
        "Profundidad en CM",
        "Nombre de variante #1",
        "Opción de variante #1",
        "Nombre de variante #2",
        "Opción de variante #2",
        "Nombre de variante #3",
        "Opción de variante #3",
        "Categorías > Subcategorías > … > Subcategorías",
    ]
    # Buscar fila por nombre
    found = None
    for r_idx in range(2, ws.max_row + 1):
        if ws.cell(row=r_idx, column=2).value == "Canon X":
            found = r_idx
            break
    assert found is not None
    # SKU canónico
    assert ws.cell(row=found, column=1).value == "AAA_0001_BBB"
    # Precio canónico
    assert float(ws.cell(row=found, column=3).value) == 99.99
    # Stock
    assert int(ws.cell(row=found, column=5).value) == 7
    # Visibilidad
    assert ws.cell(row=found, column=6).value == "Visible"
