"""Endpoints para importar listas de precios de proveedores."""

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from datetime import datetime, date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.comments import Comment

from db.models import (
    ImportJob,
    ImportJobRow,
    Supplier,
    Category,
    Product,
    SupplierProduct,
    SupplierPriceHistory,
    CanonicalProduct,
    ProductEquivalence,
)
from db.session import get_session
from services.suppliers.parsers import (
    SUPPLIER_PARSERS,
    AUTO_CREATE_CANONICAL,
    suggest_canonicals,
)
from services.auth import (
    require_roles,
    require_csrf,
    SessionData,
)

router = APIRouter()


async def _get_or_create_category_path(db: AsyncSession, path: str) -> Category:
    """Crea la jerarquía de categorías completa y devuelve la última."""
    parent_id: int | None = None
    parent: Category | None = None
    for name in [p.strip() for p in path.split(">") if p.strip()]:
        stmt = select(Category).where(
            Category.name == name, Category.parent_id == parent_id
        )
        res = await db.execute(stmt)
        cat = res.scalar_one_or_none()
        if not cat:
            cat = Category(name=name, parent_id=parent_id)
            db.add(cat)
            await db.flush()
        parent_id = cat.id
        parent = cat
    if not parent:
        raise ValueError("category_path vacío")
    return parent


async def _upsert_product(
    db: AsyncSession, code: str, title: str, category: Category
) -> Product:
    res = await db.execute(select(Product).where(Product.sku_root == code))
    prod = res.scalar_one_or_none()
    if not prod:
        prod = Product(sku_root=code, title=title, category_id=category.id)
        db.add(prod)
        await db.flush()
    else:
        prod.title = title
        prod.category_id = category.id
    return prod


async def _upsert_supplier_product(
    db: AsyncSession, supplier_id: int, data: dict, product: Product
) -> None:
    code = data["codigo"]
    parts = [p.strip() for p in data.get("categoria_path", "").split(">") if p.strip()]
    stmt = select(SupplierProduct).where(
        SupplierProduct.supplier_id == supplier_id,
        SupplierProduct.supplier_product_id == code,
    )
    res = await db.execute(stmt)
    sp = res.scalar_one_or_none()
    if not sp:
        sp = SupplierProduct(
            supplier_id=supplier_id,
            supplier_product_id=code,
        )
        db.add(sp)
    sp.title = data["nombre"]
    sp.category_level_1 = parts[0] if len(parts) > 0 else None
    sp.category_level_2 = parts[1] if len(parts) > 1 else None
    sp.category_level_3 = parts[2] if len(parts) > 2 else None
    sp.min_purchase_qty = data.get("compra_minima")
    sp.current_purchase_price = data.get("precio_compra")
    sp.current_sale_price = data.get("precio_venta")
    sp.last_seen_at = datetime.utcnow()
    sp.internal_product_id = product.id
    await db.flush()


@router.get(
    "/suppliers/price-list/template",
    dependencies=[
        Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))
    ],
)
async def download_generic_price_list_template() -> StreamingResponse:
    """Genera y descarga una plantilla Excel genérica para listas de precios."""
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    headers = [
        "ID",
        "Agrupamiento",
        "Familia",
        "SubFamilia",
        "Producto",
        "Compra Minima",
        "Stock",
        "PrecioDeCompra",
        "PrecioDeVenta",
    ]
    ws.append(headers)
    ws.append(["123", "", "", "", "Ejemplo", 1, 0, 0.0, 0.0])
    ws["A1"].comment = Comment(
        "No borres la fila de encabezados. Completa tus productos desde la fila 2.",
        "growen",
    )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    headers_resp = {
        "Content-Disposition": 'attachment; filename="plantilla-generica.xlsx"'
    }
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )


@router.get(
    "/suppliers/{supplier_id}/price-list/template",
    dependencies=[
        Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))
    ],
)
async def download_price_list_template(
    supplier_id: int, db: AsyncSession = Depends(get_session)
):
    """Genera y descarga una plantilla Excel para listas de precios."""
    supplier = await db.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    parser = SUPPLIER_PARSERS.get(supplier.slug)
    if not parser:
        raise HTTPException(
            status_code=400,
            detail=f"Proveedor no soportado: {supplier.slug}",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    headers = [
        "ID",
        "Agrupamiento",
        "Familia",
        "SubFamilia",
        "Producto",
        "Compra Minima",
        "Stock",
        "PrecioDeCompra",
        "PrecioDeVenta",
    ]
    ws.append(headers)
    ws.append(["123", "", "", "", "Ejemplo", 1, 0, 0.0, 0.0])
    ws["A1"].comment = Comment(
        "No borres la fila de encabezados. Completa tus productos desde la fila 2.",
        "growen",
    )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"plantilla-{supplier.slug}.xlsx"
    headers_resp = {
        "Content-Disposition": f"attachment; filename=\"{filename}\""
    }
    return StreamingResponse(
        stream,
        media_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )


@router.post(
    "/suppliers/{supplier_id}/price-list/upload",
    dependencies=[Depends(require_csrf)],
)
async def upload_price_list(
    supplier_id: int,
    file: UploadFile | None = File(None),
    dry_run: bool = Form(True),
    sess: SessionData = Depends(require_roles("proveedor", "colaborador", "admin")),
    db: AsyncSession = Depends(get_session),
):
    if sess.role == "proveedor" and (
        not sess.user or sess.user.supplier_id != supplier_id
    ):
        raise HTTPException(status_code=403, detail="No autorizado para este proveedor")
    if file is None:
        raise HTTPException(status_code=400, detail="file field is required")

    supplier = await db.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    parser = SUPPLIER_PARSERS.get(supplier.slug)
    if not parser:
        raise HTTPException(
            status_code=400,
            detail=f"Proveedor no soportado: {supplier.slug}",
        )

    filename = (file.filename or "").lower()
    # Admitimos tanto planillas Excel como archivos CSV
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Tipo de archivo no soportado")
    content = await file.read()

    try:
        parsed_rows = parser.parse_bytes(content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception:
        # Mensaje genérico ya que puede tratarse de Excel o CSV
        raise HTTPException(status_code=400, detail="Archivo de precios no válido")

    parser_kpis = {
        "total": len(parsed_rows),
        "errors": sum(1 for r in parsed_rows if r.get("status") == "error"),
    }

    job = ImportJob(
        supplier_id=supplier_id, filename=file.filename or "", status="DRY_RUN"
    )
    db.add(job)
    await db.flush()

    # Preparar estructuras para deduplicar y comparar contra la base
    seen_codes: set[str] = set()
    codes = [r.get("codigo") for r in parsed_rows if r.get("codigo")]
    existing_map: dict[str, SupplierProduct] = {}
    if codes:
        res = await db.execute(
            select(SupplierProduct).where(
                SupplierProduct.supplier_id == supplier_id,
                SupplierProduct.supplier_product_id.in_(codes),
            )
        )
        existing_map = {sp.supplier_product_id: sp for sp in res.scalars()}

    # Mapa de productos canónicos existentes para sugerencias por similitud
    res = await db.execute(select(CanonicalProduct.id, CanonicalProduct.name))
    canonical_map = {cid: name for cid, name in res.all()}

    kpis = {
        **parser_kpis,
        "duplicates_in_file": 0,
        "unchanged": 0,
        "new": 0,
        "changed": 0,
    }

    for idx, data in enumerate(parsed_rows):
        status = data.pop("status", "ok")
        error = data.pop("error_msg", None)

        code = data.get("codigo")

        if status != "ok":
            db.add(
                ImportJobRow(
                    job_id=job.id,
                    row_index=int(idx),
                    status=status,
                    error=error,
                    row_json_normalized=data,
                )
            )
            continue

        if not code or code in seen_codes:
            status = "duplicate_in_file"
            error = "Código duplicado en archivo"
            kpis["duplicates_in_file"] += 1
        else:
            seen_codes.add(code)
            sp = existing_map.get(code)
            if sp:
                prev_p = float(sp.current_purchase_price or 0)
                prev_s = float(sp.current_sale_price or 0)
                delta_p = round(float(data["precio_compra"]) - prev_p, 2)
                delta_s = round(float(data["precio_venta"]) - prev_s, 2)
                data["delta_compra"] = delta_p
                data["delta_venta"] = delta_s
                data["delta_pct"] = (
                    round(delta_s / prev_s * 100, 2) if prev_s else None
                )
                if delta_p == 0 and delta_s == 0:
                    status = "unchanged"
                    kpis["unchanged"] += 1
                else:
                    status = "changed"
                    kpis["changed"] += 1
            else:
                status = "new"
                kpis["new"] += 1

        # Sugerir canónico por similitud y marcar creación automática
        suggestions = suggest_canonicals(data["nombre"], canonical_map)
        data["canonical_suggestions"] = suggestions
        if AUTO_CREATE_CANONICAL and not suggestions:
            data["auto_create_canonical"] = True

        db.add(
            ImportJobRow(
                job_id=job.id,
                row_index=int(idx),
                status=status,
                error=error,
                row_json_normalized=data,
            )
        )

    job.summary_json = kpis

    await db.commit()
    return {"job_id": job.id, "summary": kpis, "kpis": kpis}


@router.get(
    "/imports/{job_id}/preview",
    dependencies=[
        Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))
    ],
)
async def preview_import(
    job_id: int,
    status: str | None = Query(None, description="Estados separados por coma"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_session),
):
    """Lista filas paginadas filtradas por ``status``."""
    res = await db.execute(select(ImportJob).where(ImportJob.id == job_id))
    job = res.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")

    stmt = select(ImportJobRow).where(ImportJobRow.job_id == job_id)
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        if statuses:
            stmt = stmt.where(ImportJobRow.status.in_(statuses))
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0
    pages = (total + page_size - 1) // page_size
    if pages and page > pages:
        page = pages

    stmt = (
        stmt.order_by(ImportJobRow.row_index)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    res = await db.execute(stmt)
    items = [
        {
            "row_index": r.row_index,
            "status": r.status,
            "error": r.error,
            "data": r.row_json_normalized,
        }
        for r in res.scalars()
    ]
    return {
        "items": items,
        "summary": job.summary_json,
        "total": total,
        "pages": pages,
        "page": page,
    }


@router.get(
    "/imports/{job_id}",
    dependencies=[
        Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))
    ],
)
async def get_import(
    job_id: int,
    limit: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_session),
):
    """Devuelve el resumen del job y las primeras ``limit`` filas."""
    res = await db.execute(select(ImportJob).where(ImportJob.id == job_id))
    job = res.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    res = await db.execute(
        select(ImportJobRow)
        .where(ImportJobRow.job_id == job_id)
        .order_by(ImportJobRow.row_index)
        .limit(limit)
    )
    rows = [
        {
            "row_index": r.row_index,
            "status": r.status,
            "error": r.error,
            "data": r.row_json_normalized,
        }
        for r in res.scalars()
    ]
    return {"job_id": job.id, "status": job.status, "summary": job.summary_json, "rows": rows}


@router.post("/imports/{job_id}/commit", dependencies=[Depends(require_csrf)])
async def commit_import(
    job_id: int,
    db: AsyncSession = Depends(get_session),
    sess: SessionData = Depends(require_roles("proveedor", "colaborador", "admin")),
):
    res = await db.execute(select(ImportJob).where(ImportJob.id == job_id))
    job = res.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    if job.status != "DRY_RUN":
        raise HTTPException(status_code=400, detail="Job ya aplicado")
    if sess.role == "proveedor" and sess.user and job.supplier_id != sess.user.supplier_id:
        raise HTTPException(status_code=403, detail="No autorizado para este proveedor")

    res = await db.execute(select(ImportJobRow).where(ImportJobRow.job_id == job_id))
    rows = res.scalars().all()

    result = {
        "inserted": 0,
        "updated": 0,
        "unchanged": 0,
        "skipped_duplicates": 0,
        "errors": 0,
        "price_changes": 0,
    }

    for r in rows:
        data = r.row_json_normalized
        if r.status == "error":
            result["errors"] += 1
            continue
        if r.status == "duplicate_in_file":
            result["skipped_duplicates"] += 1
            continue
        if r.status == "unchanged":
            result["unchanged"] += 1
            continue

        cat = await _get_or_create_category_path(db, data["categoria_path"])
        prod = await _upsert_product(db, data["codigo"], data["nombre"], cat)

        stmt = select(SupplierProduct).where(
            SupplierProduct.supplier_id == job.supplier_id,
            SupplierProduct.supplier_product_id == data["codigo"],
        )
        sp = (await db.execute(stmt)).scalar_one_or_none()
        if not sp:
            sp = SupplierProduct(
                supplier_id=job.supplier_id,
                supplier_product_id=data["codigo"],
            )
            db.add(sp)
            result["inserted"] += 1
        else:
            result["updated"] += 1

        parts = [
            p.strip() for p in data.get("categoria_path", "").split(">") if p.strip()
        ]
        sp.title = data["nombre"]
        sp.category_level_1 = parts[0] if len(parts) > 0 else None
        sp.category_level_2 = parts[1] if len(parts) > 1 else None
        sp.category_level_3 = parts[2] if len(parts) > 2 else None
        sp.min_purchase_qty = data.get("compra_minima")
        sp.last_seen_at = datetime.utcnow()
        sp.current_purchase_price = data.get("precio_compra")
        sp.current_sale_price = data.get("precio_venta")
        sp.internal_product_id = prod.id

        # Enlazar con canónicos existentes o crear nuevos según configuración
        suggestions = data.get("canonical_suggestions", [])
        if suggestions:
            canonical_id = suggestions[0]["id"]
            stmt_eq = select(ProductEquivalence).where(
                ProductEquivalence.supplier_id == job.supplier_id,
                ProductEquivalence.supplier_product_id == sp.id,
            )
            eq = (await db.execute(stmt_eq)).scalar_one_or_none()
            if eq:
                eq.canonical_product_id = canonical_id
                eq.source = "auto"
                eq.confidence = suggestions[0]["score"]
            else:
                db.add(
                    ProductEquivalence(
                        supplier_id=job.supplier_id,
                        supplier_product_id=sp.id,
                        canonical_product_id=canonical_id,
                        source="auto",
                        confidence=suggestions[0]["score"],
                    )
                )
        elif data.get("auto_create_canonical") and AUTO_CREATE_CANONICAL:
            cp = CanonicalProduct(name=data["nombre"])
            db.add(cp)
            await db.flush()
            cp.ng_sku = f"NG-{cp.id:06d}"
            db.add(
                ProductEquivalence(
                    supplier_id=job.supplier_id,
                    supplier_product_id=sp.id,
                    canonical_product_id=cp.id,
                    source="auto",
                    confidence=1.0,
                )
            )

        if r.status == "changed":
            prev_purchase = data.get("precio_compra") - data.get("delta_compra", 0)
            prev_sale = data.get("precio_venta") - data.get("delta_venta", 0)
            delta_purchase_pct = (
                (data.get("delta_compra", 0) / prev_purchase * 100)
                if prev_purchase
                else None
            )
            delta_sale_pct = (
                (data.get("delta_venta", 0) / prev_sale * 100)
                if prev_sale
                else None
            )
            db.add(
                SupplierPriceHistory(
                    supplier_product_fk=sp.id,
                    file_fk=None,
                    as_of_date=date.today(),
                    purchase_price=data.get("precio_compra"),
                    sale_price=data.get("precio_venta"),
                    delta_purchase_pct=delta_purchase_pct,
                    delta_sale_pct=delta_sale_pct,
                )
            )
            result["price_changes"] += 1

    job.status = "COMMITTED"
    await db.commit()
    return result
