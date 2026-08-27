#!/usr/bin/env python
# NG-HEADER: Nombre de archivo: catalog.py
# NG-HEADER: Ubicación: services/routers/catalog.py
# NG-HEADER: Descripción: Endpoints de catálogo (productos mínimos, proveedores, archivos, categorías, etc.)
# NG-HEADER: Lineamientos: Ver AGENTS.md
"""Endpoints para gestionar proveedores y categorías."""
from __future__ import annotations

import os
import csv
from decimal import Decimal
from enum import Enum
from typing import List, Optional, Literal
import hashlib
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File, Form
from datetime import datetime as _dt
from fastapi.responses import JSONResponse, StreamingResponse
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel, ValidationError, Field, field_validator
from sqlalchemy import func, select, or_, and_, update, exists
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
import logging

from db.models import (
    Category,
    Supplier,
    SupplierFile,
    SupplierPriceHistory,
    SupplierProduct,
    Product,
    Image,
    Variant,
    Inventory,
    ProductEquivalence,
    CanonicalProduct,
    AuditLog,
    Purchase,
    PurchaseLine,
    PurchaseAttachment,
    StockLedger,
    Tag,
    ProductTag,
    CanonicalEnrichmentJob,
    CanonicalContentVersion,
)
from db.session import get_session
from db.text_utils import stylize_product_name
from agent_core.config import settings
from ai.router import AIRouter
from ai.providers.openai_provider import OpenAIProvider
from ai.types import Task
from agent_core.detect_mcp_url import get_mcp_web_search_url
from services.auth import require_csrf, require_roles, current_session, SessionData

logger = logging.getLogger(__name__)

router = APIRouter(tags=["catalog"])

# Tamaño de página por defecto para el historial de precios
DEFAULT_PRICE_HISTORY_PAGE_SIZE = int(os.getenv("PRICE_HISTORY_PAGE_SIZE", "20"))


@router.get("/products/{product_id}/purchase-history", dependencies=[Depends(require_roles("colaborador", "admin"))])
async def product_purchase_history(product_id: int, session: AsyncSession = Depends(get_session)):
    """Historial confirmado unido para todos los internos del mismo canónico."""
    product = await session.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    canonical_id = await session.scalar(
        select(ProductEquivalence.canonical_product_id)
        .join(SupplierProduct, SupplierProduct.id == ProductEquivalence.supplier_product_id)
        .where(SupplierProduct.internal_product_id == product_id)
        .limit(1)
    )
    linked_ids = [product_id]
    if canonical_id:
        linked_ids = list(
            (
                await session.scalars(
                    select(SupplierProduct.internal_product_id)
                    .join(ProductEquivalence, ProductEquivalence.supplier_product_id == SupplierProduct.id)
                    .where(
                        ProductEquivalence.canonical_product_id == canonical_id,
                        SupplierProduct.internal_product_id.is_not(None),
                    )
                    .distinct()
                )
            ).all()
        ) or linked_ids
    rows = (
        await session.execute(
            select(PurchaseLine, Purchase, Supplier)
            .join(Purchase, Purchase.id == PurchaseLine.purchase_id)
            .join(Supplier, Supplier.id == Purchase.supplier_id)
            .where(PurchaseLine.product_id.in_(linked_ids), Purchase.status == "CONFIRMADA")
            .order_by(Purchase.remito_date.desc(), PurchaseLine.id.desc())
        )
    ).all()
    items = []
    for line, purchase, supplier in rows:
        attachment = await session.scalar(
            select(PurchaseAttachment)
            .where(PurchaseAttachment.purchase_id == purchase.id)
            .order_by(PurchaseAttachment.id.asc())
        )
        discount = float(line.line_discount or 0)
        gross = float(line.unit_cost or 0)
        items.append({
            "purchase_id": purchase.id,
            "purchase_line_id": line.id,
            "product_id": line.product_id,
            "date": purchase.remito_date.isoformat(),
            "supplier": {"id": supplier.id, "name": supplier.name},
            "remito_number": purchase.remito_number,
            "supplier_sku": line.supplier_sku,
            "supplier_title": line.title,
            "quantity": int(line.qty or 0),
            "gross_unit_cost": gross,
            "discount_pct": discount,
            "net_unit_cost": round(gross * (1 - discount / 100), 2),
            "attachment_url": f"/purchases/{purchase.id}/attachments/{attachment.id}/file" if attachment else None,
        })
    movements = (
        await session.execute(
            select(StockLedger)
            .where(
                StockLedger.product_id.in_(linked_ids),
                StockLedger.source_type.in_(("purchase", "purchase_rollback")),
            )
            .order_by(StockLedger.created_at.desc())
        )
    ).scalars().all()
    return {
        "product_id": product.id,
        "product_name": product.title,
        "items": items,
        "movements": [{
            "type": movement.source_type,
            "product_id": movement.product_id,
            "source_id": movement.source_id,
            "delta": movement.delta,
            "balance_after": movement.balance_after,
            "created_at": movement.created_at.isoformat(),
        } for movement in movements],
    }
# ------------------------------- Productos (mínimo para tests) -------------------------------
from pydantic import BaseModel as _PydModel


class _ProductCreate(_PydModel):
    title: str
    initial_stock: Decimal = Decimal("0")
    supplier_id: Optional[int] = None
    supplier_sku: Optional[str] = None
    sku: Optional[str] = None
    purchase_price: Optional[float] = None
    sale_price: Optional[float] = None
    # Nuevos campos para generación canónica automática
    category_name: Optional[str] = None
    subcategory_name: Optional[str] = None
    generate_canonical: bool = False


_PRODUCTS_HAS_CANONICAL_COL: bool | None = None


async def _products_has_canonical(session: AsyncSession) -> bool:
    """Detecta (y si es posible crea en caliente para SQLite) la columna canonical_sku.

    Evita depender exclusivamente de migraciones en entorno de tests que ya
    tenían la tabla creada antes de introducir el campo en el modelo.
    """
    global _PRODUCTS_HAS_CANONICAL_COL
    if _PRODUCTS_HAS_CANONICAL_COL is not None:
        return _PRODUCTS_HAS_CANONICAL_COL
    try:
        bind = session.get_bind()
        dialect = bind.dialect.name if bind else ""
        if dialect == "sqlite":
            res = await session.execute("PRAGMA table_info(products)")  # type: ignore[arg-type]
            cols = [row[1] for row in res.fetchall()]  # row[1] = name
            if "canonical_sku" in cols:
                _PRODUCTS_HAS_CANONICAL_COL = True
                return True
            # Intentar agregar columna en caliente (primer intento dentro de la sesión)
            try:
                await session.execute("ALTER TABLE products ADD COLUMN canonical_sku VARCHAR(32)")  # type: ignore[arg-type]
                await session.commit()
            except Exception as e:
                # Reintento usando conexión en modo autocommit (algunos entornos SQLite pueden requerirlo)
                try:
                    await session.rollback()
                except Exception:
                    pass
                try:
                    bind_conn = session.get_bind()
                    if bind_conn is not None:
                        await bind_conn.execution_options(isolation_level="AUTOCOMMIT").execute("ALTER TABLE products ADD COLUMN canonical_sku VARCHAR(32)")  # type: ignore[arg-type]
                except Exception:
                    _PRODUCTS_HAS_CANONICAL_COL = False
                    return False
            # Verificar nuevamente
            res2 = await session.execute("PRAGMA table_info(products)")  # type: ignore[arg-type]
            cols2 = [row[1] for row in res2.fetchall()]
            if "canonical_sku" in cols2:
                _PRODUCTS_HAS_CANONICAL_COL = True
                return True
            _PRODUCTS_HAS_CANONICAL_COL = False
            return False
        else:
            q = """
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'products' AND column_name = 'canonical_sku'
            LIMIT 1
            """
            try:
                res = await session.execute(q)  # type: ignore[arg-type]
                if res.first():
                    _PRODUCTS_HAS_CANONICAL_COL = True
                else:
                    _PRODUCTS_HAS_CANONICAL_COL = False
            except Exception:
                _PRODUCTS_HAS_CANONICAL_COL = False
        return _PRODUCTS_HAS_CANONICAL_COL
    except Exception:
        _PRODUCTS_HAS_CANONICAL_COL = False
        return False


@router.post(
    "/catalog/products",
    dependencies=[Depends(require_csrf)],
)
async def create_product_minimal(payload: _ProductCreate, session: AsyncSession = Depends(get_session)):
    """Crea un producto mínimo para pruebas con un Variant y opcionalmente inventario.

    Nota: endpoint pensado para entorno de pruebas; en producción existen flujos más ricos.
    """
    supplier = None
    if payload.supplier_id is not None:
        supplier = await session.get(Supplier, payload.supplier_id)
        if not supplier:
            raise HTTPException(status_code=400, detail={"code": "invalid_supplier_id", "message": "supplier_id inválido"})

    from db.models import Variant, Inventory, SupplierProduct, SupplierPriceHistory
    desired_sku = (payload.sku or payload.supplier_sku or payload.title)[:50].strip() if payload.sku or payload.supplier_sku else (payload.title or "")[:50].strip()
    if not desired_sku:
        raise HTTPException(status_code=400, detail={"code": "invalid_sku", "message": "SKU inválido"})
    from db.sku_utils import is_canonical_sku, CANONICAL_SKU_PATTERN, CANONICAL_SKU_REGEX
    strict_flag = os.getenv("CANONICAL_SKU_STRICT", "1") == "1"  # ahora estricto por defecto
    force_gen_flag = os.getenv("FORCE_CANONICAL", "0") == "1"

    # Regla pseudo-canónica: si tiene exactamente dos '_' y no cumple regex => 422
    if desired_sku.count('_') == 2 and not is_canonical_sku(desired_sku):
        raise HTTPException(status_code=422, detail={
            "code": "invalid_canonical_sku",
            "message": f"Formato canónico inválido: esperado {CANONICAL_SKU_PATTERN}",
        })
    # Adicional: si parece canónico pero sin guiones bajos (AAA0000BBB), rechazar
    try:
        import re as _re
        if _re.match(r"^[A-Za-z]{3}\d{4}[A-Za-z]{3}$", desired_sku or "") and not is_canonical_sku(desired_sku):
            raise HTTPException(status_code=422, detail={
                "code": "invalid_canonical_sku",
                "message": f"Formato canónico inválido: esperado {CANONICAL_SKU_PATTERN}",
            })
    except HTTPException:
        raise

    sku_is_canonical = is_canonical_sku(desired_sku)

    # Si el SKU ya existe y es canónico, permitir vincular SupplierProduct en lugar de error (linking)
    if sku_is_canonical:
        existing_var = await session.scalar(select(Variant).where(func.lower(Variant.sku) == desired_sku.lower()))
        if existing_var:
            existing_prod = await session.get(Product, existing_var.product_id)
            if payload.supplier_id is not None:
                # Si ya existe un SupplierProduct con la misma pareja (supplier_id, supplier_sku),
                # consideramos que es un duplicado de SKU del flujo minimal y devolvemos 409 duplicate_sku
                if payload.supplier_sku:
                    sp_exist = await session.scalar(
                        select(SupplierProduct).where(
                            SupplierProduct.supplier_id == payload.supplier_id,
                            SupplierProduct.supplier_product_id == payload.supplier_sku,
                        )
                    )
                    if sp_exist:
                        raise HTTPException(status_code=409, detail={"code": "duplicate_sku", "message": "SKU ya existente"})
                # Crear/asegurar SupplierProduct vinculado al producto/variante existente
                sp = SupplierProduct(
                    supplier_id=payload.supplier_id,
                    supplier_product_id=(payload.supplier_sku or desired_sku),
                    title=payload.title[:200],
                    current_purchase_price=(payload.purchase_price if payload.purchase_price is not None else None),
                    current_sale_price=(payload.sale_price if payload.sale_price is not None else None),
                    internal_product_id=existing_prod.id if existing_prod else None,
                    internal_variant_id=existing_var.id,
                )
                session.add(sp)
                await session.flush()
                try:
                    await session.commit()
                except IntegrityError:
                    await session.rollback()
                    # Si ya existe SupplierProduct con ese supplier+sku, retornamos link sin crear
                response = {
                    "id": existing_prod.id if existing_prod else None,
                    "title": existing_prod.title if existing_prod else payload.title,
                    "sku_root": getattr(existing_prod, 'sku_root', desired_sku) if existing_prod else desired_sku,
                    "linked": True,
                    "created": False,
                    "idempotent": False,
                }
                try:
                    response["supplier_item_id"] = sp.id
                except Exception:
                    pass
                return response
            # Sin supplier, retornar referencia sin crear duplicados
            return {
                "id": existing_prod.id if existing_prod else None,
                "title": existing_prod.title if existing_prod else payload.title,
                "sku_root": getattr(existing_prod, 'sku_root', desired_sku) if existing_prod else desired_sku,
                "linked": True,
                "created": False,
                "idempotent": True,
            }

    # Generación automática si se solicita o es requerido en modo estricto sin sku válido
    if (payload.generate_canonical or (strict_flag and not sku_is_canonical)):
        # Requiere category_name y subcategory_name (subcat opcional, si falta se reutiliza category)
        if not payload.category_name:
            raise HTTPException(status_code=400, detail={"code": "missing_category_name", "message": "category_name requerido para generación canónica"})
        from db.sku_generator import generate_canonical_sku, CanonicalSkuGenerationError
        try:
            desired_sku = await generate_canonical_sku(session, payload.category_name, payload.subcategory_name or payload.category_name)
            sku_is_canonical = True
        except CanonicalSkuGenerationError as ge:
            raise HTTPException(status_code=500, detail={"code": "canonical_generation_error", "message": str(ge)})

    # En modo no estricto, aceptamos legacy y sólo seteamos canonical_sku si coincide el patrón.
    # Búsqueda case-insensitive para evitar conflictos por mayúsculas/minúsculas
    existing = await session.scalar(select(Variant).where(func.lower(Variant.sku) == desired_sku.lower()))
    if existing:
        raise HTTPException(status_code=409, detail={"code": "duplicate_sku", "message": "SKU ya existente"})

    try:
        # Asegurar (o crear en caliente en SQLite) la columna canonical_sku ANTES de instanciar Product
        has_canonical_col = await _products_has_canonical(session)
        import logging as _logging
        _logging.getLogger("growen").debug({"event": "create_product_minimal.start", "desired_sku": desired_sku, "strict": strict_flag})

        # Crear producto con canonical_sku si el SKU es canónico
        prod = Product(sku_root=desired_sku, title=payload.title)
        if sku_is_canonical and has_canonical_col:
            prod.canonical_sku = desired_sku
        session.add(prod)
        await session.flush()
        # Si el SKU elegido ya se usó (race o fallback previo) y estamos en modo no estricto, generar sufijo incremental
        attempt_sku = desired_sku
        attempt_idx = 1
        while True:
            conflict = await session.scalar(select(Variant).where(func.lower(Variant.sku) == attempt_sku.lower()))
            if not conflict:
                break
            if strict_flag:
                raise HTTPException(status_code=409, detail={"code": "duplicate_sku", "message": "SKU ya existente"})
            attempt_idx += 1
            suffix = f"-{attempt_idx}" if attempt_idx < 10 else f"-{attempt_idx}"
            base_len = 50 - len(suffix)
            attempt_sku = (desired_sku[:base_len] + suffix)[:50]
            _logging.getLogger("growen").debug({"event": "create_product_minimal.retry_sku", "attempt": attempt_idx, "attempt_sku": attempt_sku})
            if attempt_idx > 25:
                raise HTTPException(status_code=409, detail={"code": "duplicate_sku", "message": "No se pudo generar SKU único"})
        if attempt_sku != desired_sku:
            desired_sku = attempt_sku
            # Actualizar sku_root y canonical_sku si aplica
            prod.sku_root = desired_sku
            if is_canonical_sku(desired_sku) and has_canonical_col:
                prod.canonical_sku = desired_sku
        # Crear Variant con reintentos en caso de colisión de unicidad (modo no estricto)
        import random as _r, string as _s
        max_variant_retries = 6
        last_error = None
        var = None
        for vr in range(max_variant_retries):
            attempt_variant_sku = desired_sku if vr == 0 else (
                (desired_sku[:40] + "-" + ''.join(_r.choices(_s.ascii_uppercase + _s.digits, k=5)))[:50]
            )
            try:
                var = Variant(product_id=prod.id, sku=attempt_variant_sku)
                session.add(var)
                await session.flush()
                if attempt_variant_sku != desired_sku:
                    desired_sku = attempt_variant_sku
                    prod.sku_root = desired_sku
                    if is_canonical_sku(desired_sku) and has_canonical_col:
                        prod.canonical_sku = desired_sku
                break
            except IntegrityError as ie:  # collision
                last_error = ie
                await session.rollback()
                # Reanudar transacción lógica: necesitamos asegurar que prod sigue presente (en stub path ya está)
                # Reiniciar sesión para siguiente intento
                # Nota: rollback no elimina el INSERT manual previo.
                if strict_flag:
                    raise HTTPException(status_code=409, detail={"code": "duplicate_sku", "message": "SKU ya existente"})
                continue
        if var is None:
            # No se pudo generar SKU único tras reintentos
            raise HTTPException(status_code=409, detail={"code": "duplicate_sku", "message": "Colisión repetida en SKU"})
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=409, detail={"code": "duplicate_sku", "message": "SKU ya existente"})
    # Compatibilidad: reflejar product_id/variant_id directos en Product para flujos que esperan product.stock
    try:
        if getattr(prod, 'stock', None) is None:
            prod.stock = 0  # aseguramos campo
    except Exception:
        pass

    # Inventario opcional
    if payload.initial_stock and payload.initial_stock > 0:
        inv = Inventory(variant_id=var.id, stock_qty=payload.initial_stock)
        session.add(inv)

    # Guardar stock agregado también en Product.stock para compatibilidad
    prod.stock = Decimal(str(payload.initial_stock or 0))

    # Crear SupplierProduct asociado si hay supplier_id
    if supplier is not None:
        sp = SupplierProduct(
            supplier_id=payload.supplier_id,
            supplier_product_id=(payload.supplier_sku or prod.sku_root),
            title=payload.title[:200],
            current_purchase_price=(payload.purchase_price if payload.purchase_price is not None else None),
            current_sale_price=(payload.sale_price if payload.sale_price is not None else None),
            internal_product_id=prod.id,
            internal_variant_id=var.id,
        )
        session.add(sp)
        await session.flush()

        # Si se envió purchase_price pero no sale_price, por defecto igualar venta a compra
        if payload.purchase_price is not None and payload.sale_price is None:
            try:
                sp.current_sale_price = payload.purchase_price
            except Exception:
                pass

        # Registrar historial de precios solo si se enviaron ambos precios
        if payload.purchase_price is not None and payload.sale_price is not None:
            from datetime import date as _date
            sph = SupplierPriceHistory(
                supplier_product_fk=sp.id,
                file_fk=None,
                as_of_date=_date.today(),
                purchase_price=payload.purchase_price,
                sale_price=payload.sale_price,
                delta_purchase_pct=None,
                delta_sale_pct=None,
            )
            session.add(sph)

    try:
        await session.commit()
        # Priorizar canonical_sku en la respuesta
        response_sku = prod.canonical_sku or prod.sku_root
        response = {"id": prod.id, "title": prod.title, "sku_root": response_sku, "canonical_sku": prod.canonical_sku, "idempotent": False, "created": True}
    except IntegrityError:
        await session.rollback()
        if not strict_flag:
            # Buscar variant existente por sku_root (case-insensitive)
            v_exist = await session.scalar(select(Variant).where(func.lower(Variant.sku) == desired_sku.lower()))
            if v_exist:
                p_exist = await session.get(Product, v_exist.product_id)
                return {"id": p_exist.id if p_exist else None, "title": p_exist.title if p_exist else payload.title, "sku_root": getattr(p_exist, 'sku_root', desired_sku), "idempotent": True, "created": False}
        raise
    if supplier is not None:
        # sp puede existir si creamos SupplierProduct
        try:  # defensivo en caso de refactors
            response["supplier_item_id"] = sp.id  # type: ignore[name-defined]
        except NameError:
            pass
    return response


# ------------------------------- Proveedores: búsqueda (autocomplete) -------------------------------
class _SupplierSearchItem(_PydModel):
    id: int
    name: str
    slug: str


@router.get(
    "/suppliers/search",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def suppliers_search(
    q: str = Query("", description="Texto a buscar en name|slug"),
    limit: int = Query(20, ge=1, le=50),
    session: AsyncSession = Depends(get_session),
):
    """Autocomplete de proveedores por name|slug. Retorna hasta `limit` elementos ordenados por nombre.

    Cuando `q` viene vac��o se devuelve el top `limit` ordenado por nombre (uso como combo-box inicial).
    """
    q_clean = (q or "").strip()

    stmt = (
        select(Supplier.id, Supplier.name, Supplier.slug)
        .order_by(Supplier.name.asc())
        .limit(limit)
    )
    if q_clean:
        pattern = f"%{q_clean}%"
        stmt = stmt.where(or_(Supplier.name.ilike(pattern), Supplier.slug.ilike(pattern)))

    rows = (await session.execute(stmt)).all()
    return [{"id": r[0], "name": r[1], "slug": r[2]} for r in rows]


# ------------------------------- Variants: editar SKU interno -------------------------------
class _VariantSkuUpdate(_PydModel):
    sku: str
    note: Optional[str] = None


@router.put(
    "/variants/{variant_id}/sku",
    dependencies=[Depends(require_csrf), Depends(require_roles("colaborador", "admin"))],
)
async def update_variant_sku(
    variant_id: int,
    payload: _VariantSkuUpdate,
    request: Request,
    session_data: SessionData = Depends(current_session),
    session: AsyncSession = Depends(get_session),
):
    """Actualiza el SKU interno de una variante con validación de formato y unicidad.

    - Regex permitida: [A-Za-z0-9._\\-]{2,50}
    - Unicidad global en `Variant.sku` (existe constraint de DB adicional)
    - Auditoría en `AuditLog` (action: variant.sku.update)
    """
    import re

    new_sku = payload.sku.strip()
    if not re.fullmatch(r"[A-Za-z0-9._\-]{2,50}", new_sku):
        raise HTTPException(status_code=400, detail={"code": "invalid_sku_format", "message": "Formato de SKU inválido"})

    var = await session.get(Variant, variant_id)
    if not var:
        raise HTTPException(status_code=404, detail={"code": "variant_not_found"})

    if var.sku == new_sku:
        return {"id": var.id, "sku": var.sku, "unchanged": True}

    exists = await session.scalar(select(func.count()).select_from(Variant).where(Variant.sku == new_sku))
    if (exists or 0) > 0:
        raise HTTPException(status_code=409, detail={"code": "duplicate_sku", "message": "SKU ya existente"})

    old_sku = var.sku
    var.sku = new_sku
    await session.flush()

    # Auditoría
    try:
        audit = AuditLog(
            action="variant.sku.update",
            table="variants",
            entity_id=var.id,
            meta={"old": old_sku, "new": new_sku, "note": payload.note},
            user_id=(session_data.user_id if session_data else None),
            ip=request.client.host if request and request.client else None,
        )
        session.add(audit)
    except Exception:
        pass

    await session.commit()
    return {"id": var.id, "sku": var.sku}


# ------------------------------- Búsqueda rápida de catálogo (POS) -------------------------------
class _CatalogSearchItem(_PydModel):
    id: int
    kind: str  # product|canonical
    title: str
    sku: str | None = None
    stock: int | None = None
    price: float | None = None


@router.get(
    "/catalog/search",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def catalog_search(
    q: str = Query("", description="Texto a buscar en título/SKU/descripción"),
    limit: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_session),
):
    """Búsqueda rápida para POS y chatbot.

    Busca productos con su información canónica vinculada.
    Busca en: título, SKU, descripción HTML, y términos relacionados.
    Devuelve SKU canónico (formato XXX_####_YYY) preferentemente sobre el interno.
    Prioriza productos con stock>0.
    
    Términos especiales reconocidos:
    - "vegetativo", "veg", "crecimiento" → busca en descripción
    - "nitrógeno", "N", "nitrogeno" → busca en descripción
    - "floración", "flora", "flor" → busca en descripción
    """
    term = (q or "").strip()
    
    # Subquery para imagen principal
    from db.models import Image
    primary_image_path = (
        select(Image.path)
        .where(Image.product_id == Product.id)
        .where(Image.active == True)
        .order_by(Image.is_primary.desc(), Image.sort_order.asc(), Image.id.asc())
        .limit(1)
        .scalar_subquery()
    )

    # Query base: productos con su info canónica vinculada
    base_query = (
        select(
            Product.id,
            Product.title,
            Product.stock,
            Product.description_html,
            CanonicalProduct.id.label("canonical_id"),
            CanonicalProduct.name.label("canonical_name"),
            CanonicalProduct.sku_custom.label("canonical_sku"),
            CanonicalProduct.ng_sku,
            CanonicalProduct.sale_price,
            primary_image_path.label("image_path"),
        )
        .join(SupplierProduct, SupplierProduct.internal_product_id == Product.id, isouter=True)
        .join(ProductEquivalence, ProductEquivalence.supplier_product_id == SupplierProduct.id, isouter=True)
        .join(CanonicalProduct, CanonicalProduct.id == ProductEquivalence.canonical_product_id, isouter=True)
    )
    
    if not term:
        # Top por stock (productos con stock)
        rows = (
            await session.execute(
                base_query
                .where((Product.stock != None) & (Product.stock > 0))
                .order_by(Product.stock.desc(), Product.title.asc())
                .limit(limit)
            )
        ).all()
    else:
        # Búsqueda por palabras clave (AND logic)
        terms = term.split()
        
        # Lista de condiciones AND (una por cada palabra)
        and_conditions = []
        
        for t in terms:
            w_like = f"%{t}%"
            # Para cada palabra, debe machear en al menos uno de los campos (OR)
            or_conditions = [
                Product.title.ilike(w_like),
                CanonicalProduct.name.ilike(w_like),
                CanonicalProduct.sku_custom.ilike(w_like),
                CanonicalProduct.ng_sku.ilike(w_like),
                Product.description_html.ilike(w_like),
                exists(
                    select(1)
                    .select_from(ProductTag)
                    .join(Tag, Tag.id == ProductTag.tag_id)
                    .where(ProductTag.product_id == Product.id, Tag.name.ilike(w_like))
                ),
            ]
            and_conditions.append(or_(*or_conditions))
        
        # Normalizar término completo para búsqueda de relaciones (opcional)
        term_lower = term.lower()
        related_terms = []
        
        # Mapeo de términos relacionados (se agregan como OR global o se refinan?)
        # Nota: La lógica anterior usaba OR global. Para mantener simplicidad y potencia,
        # usaremos la lógica de palabras clave AND como base.
        # Los related terms (veg, flora) podrían sumarse a la query si no hay resultados,
        # pero por ahora priorizamos la búsqueda exacta de palabras.
        
        stmt = base_query.where(and_(*and_conditions))
        
        # Ordenar: primero stock, luego título
        stmt = stmt.order_by(Product.stock.desc().nullslast(), Product.title.asc())
        stmt = stmt.limit(limit * 2)
        
        rows = (await session.execute(stmt)).all()
    
    # Deduplicar por product_id (puede haber múltiples filas si hay varios SupplierProducts)
    seen_ids: set[int] = set()
    items: list[dict] = []
    
    # Obtener todos los product_ids únicos para consultar tags en batch
    product_ids = {row.id for row in rows}
    tags_map: dict[int, list[str]] = {}
    if product_ids:
        try:
            tag_result = (
                await session.execute(
                    select(ProductTag.product_id, Tag.name)
                    .join(Tag, ProductTag.tag_id == Tag.id)
                    .where(ProductTag.product_id.in_(product_ids))
                )
            ).all()
            for product_id, tag_name in tag_result:
                if product_id not in tags_map:
                    tags_map[product_id] = []
                tags_map[product_id].append(f"#{tag_name}")
        except Exception:
            # Si falla la consulta de tags, continuar sin tags
            pass
    
    for row in rows:
        if row.id in seen_ids:
            continue
        seen_ids.add(row.id)
        
        # SKU preferido: canónico (formato XXX_####_YYY) sobre interno
        preferred_sku = row.canonical_sku or row.ng_sku
        # Nombre preferido: canónico sobre interno
        preferred_name = stylize_product_name(row.canonical_name or row.title) or row.title
        # Precio de venta desde canónico
        sale_price = float(row.sale_price) if row.sale_price else None
        # Tags del producto
        tags = tags_map.get(row.id, [])
        
        items.append({
            "id": row.id,
            "title": preferred_name,
            "sku": preferred_sku,
            "stock": int(row.stock or 0),
            "price": sale_price,
            "has_description": bool(row.description_html),
            "tags": tags,  # Lista de tags formateados como ["#Organico", "#Floracion"]
        })
    
    # Orden: productos con stock primero; luego por nombre
    items.sort(key=lambda it: (0 if (it.get("stock") or 0) > 0 else 1, (it.get("title") or "")))
    
    return items[:limit]


@router.get(
    "/catalog/search_by_tags",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def catalog_search_by_tags(
    tags: str = Query(..., description="Tags separados por coma (ej: 'Organico,Floracion')"),
    limit: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_session),
):
    """Búsqueda de productos por tags.
    
    Busca productos que tengan todos los tags especificados.
    Devuelve productos con su información canónica y tags asociados.
    """
    # Parsear tags
    tag_names = [t.strip() for t in tags.split(",") if t.strip()]
    if not tag_names:
        raise HTTPException(status_code=400, detail="Se requiere al menos un tag")
    
    # Normalizar nombres de tags (remover # si existe)
    tag_names = [t.lstrip("#") for t in tag_names]
    
    try:
        from db.models import Tag, ProductTag
        
        # Buscar productos que tengan todos los tags especificados
        # Usar subquery para cada tag y hacer intersección
        base_query = (
            select(
                Product.id,
                Product.title,
                Product.stock,
                Product.description_html,
                CanonicalProduct.id.label("canonical_id"),
                CanonicalProduct.name.label("canonical_name"),
                CanonicalProduct.sku_custom.label("canonical_sku"),
                CanonicalProduct.ng_sku,
                CanonicalProduct.sale_price,
            )
            .join(SupplierProduct, SupplierProduct.internal_product_id == Product.id, isouter=True)
            .join(ProductEquivalence, ProductEquivalence.supplier_product_id == SupplierProduct.id, isouter=True)
            .join(CanonicalProduct, CanonicalProduct.id == ProductEquivalence.canonical_product_id, isouter=True)
        )
        
        # Construir condición: productos que tengan todos los tags
        # Usar subquery para cada tag
        tag_conditions = []
        for tag_name in tag_names:
            tag_subquery = (
                select(ProductTag.product_id)
                .join(Tag, ProductTag.tag_id == Tag.id)
                .where(func.lower(Tag.name) == tag_name.lower())
            )
            tag_conditions.append(Product.id.in_(tag_subquery))
        
        # Productos que cumplen todas las condiciones (AND)
        if len(tag_conditions) == 1:
            final_condition = tag_conditions[0]
        else:
            from sqlalchemy import and_
            final_condition = and_(*tag_conditions)
        
        rows = (
            await session.execute(
                base_query
                .where(final_condition)
                .order_by(Product.stock.desc().nullslast(), Product.title.asc())
                .limit(limit * 2)  # Extra para deduplicar
            )
        ).all()
        
        # Deduplicar y obtener tags
        seen_ids: set[int] = set()
        items: list[dict] = []
        
        product_ids = {row.id for row in rows}
        tags_map: dict[int, list[str]] = {}
        if product_ids:
            tag_result = (
                await session.execute(
                    select(ProductTag.product_id, Tag.name)
                    .join(Tag, ProductTag.tag_id == Tag.id)
                    .where(ProductTag.product_id.in_(product_ids))
                )
            ).all()
            for product_id, tag_name in tag_result:
                if product_id not in tags_map:
                    tags_map[product_id] = []
                tags_map[product_id].append(f"#{tag_name}")
        
        for row in rows:
            if row.id in seen_ids:
                continue
            seen_ids.add(row.id)
            
            preferred_sku = row.canonical_sku or row.ng_sku
            preferred_name = stylize_product_name(row.canonical_name or row.title) or row.title
            sale_price = float(row.sale_price) if row.sale_price else None
            product_tags = tags_map.get(row.id, [])
            
            items.append({
                "id": row.id,
                "title": preferred_name,
                "sku": preferred_sku,
                "stock": int(row.stock or 0),
                "price": sale_price,
                "has_description": bool(row.description_html),
                "tags": product_tags,
            })
        
        # Ordenar: productos con stock primero
        items.sort(key=lambda it: (0 if (it.get("stock") or 0) > 0 else 1, (it.get("title") or "")))
        
        return {
            "items": items[:limit],
            "count": len(items[:limit]),
            "tags": tag_names,
        }
        
    except Exception as e:
        logger.error(f"Error en búsqueda por tags: {e}")
        raise HTTPException(status_code=500, detail=f"Error al buscar productos por tags: {e}")


# ------------------------------- Secuencia para SKUs Canónicos -------------------------------
@router.get(
    "/catalog/next-seq",
    dependencies=[Depends(require_roles("colaborador", "admin"))],
)
async def catalog_next_seq(
    category_id: int | None = Query(None, description="ID de categoría (opcional, si no se envía se calcula global)"),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Devuelve la próxima secuencia disponible para SKUs canónicos (XXX_####_YYY).
    
    El cálculo se basa en extraer el número máximo de secuencia de los SKUs existentes,
    no simplemente en contar productos. Esto evita colisiones cuando se eliminan productos.
    """
    import re
    
    # Consultar todos los sku_custom de la categoría (o global si no hay category_id)
    if category_id is not None:
        stmt = select(CanonicalProduct.sku_custom).where(
            CanonicalProduct.category_id == category_id,
            CanonicalProduct.sku_custom.isnot(None)
        )
    else:
        stmt = select(CanonicalProduct.sku_custom).where(
            CanonicalProduct.category_id.is_(None),
            CanonicalProduct.sku_custom.isnot(None)
        )
    
    result = await session.execute(stmt)
    skus = [row[0] for row in result.all() if row[0]]
    
    # Extraer el número de secuencia de cada SKU (formato: XXX_####_YYY)
    # El número es la parte central entre los dos underscores
    pattern = re.compile(r'^[A-Z]{3}_(\d{4})_[A-Z]{3}$')
    max_seq = 0
    
    for sku in skus:
        if sku:
            match = pattern.match(sku.upper())
            if match:
                seq_num = int(match.group(1))
                if seq_num > max_seq:
                    max_seq = seq_num
    
    # La siguiente secuencia es max + 1 (mínimo 1)
    next_seq = max_seq + 1
    
    return {
        "category_id": category_id,
        "next_seq": next_seq,
        "max_existing": max_seq,
    }


# ------------------------------- Helper: Build Product Response -------------------------------
async def _build_product_response(session: AsyncSession, product: Product) -> dict:
    """Construye la respuesta completa de un producto con su info canónica.
    
    Devuelve SKU canónico (formato XXX_####_YYY) preferentemente.
    """
    # Calcular stock real
    stock = product.stock or 0
    try:
        inv_result = await session.execute(
            select(func.sum(Inventory.stock_qty))
            .join(Variant, Variant.id == Inventory.variant_id)
            .where(Variant.product_id == product.id)
        )
        inv_total = inv_result.scalar()
        if inv_total is not None:
            stock = int(inv_total)
    except Exception:
        pass

    # Obtener info canónica vinculada
    canonical_info = (
        await session.execute(
            select(CanonicalProduct)
            .join(ProductEquivalence, ProductEquivalence.canonical_product_id == CanonicalProduct.id)
            .join(SupplierProduct, SupplierProduct.id == ProductEquivalence.supplier_product_id)
            .where(SupplierProduct.internal_product_id == product.id)
            .limit(1)
        )
    ).scalars().first()

    # SKU preferido: canónico (formato XXX_####_YYY) sobre interno
    canonical_sku = None
    sale_price = None
    canonical_name = None
    
    if canonical_info:
        canonical_sku = canonical_info.sku_custom or canonical_info.ng_sku
        canonical_name = canonical_info.name
        if canonical_info.sale_price:
            sale_price = float(canonical_info.sale_price)

    # Si no hay precio canónico, intentar desde variante
    if sale_price is None:
        variant = (
            await session.execute(
                select(Variant).where(Variant.product_id == product.id).limit(1)
            )
        ).scalars().first()
        if variant and (variant.promo_price or variant.price):
            sale_price = float(variant.promo_price or variant.price)

    # Obtener tags del producto
    tags = []
    try:
        from db.models import Tag, ProductTag
        tag_result = (
            await session.execute(
                select(Tag.name)
                .join(ProductTag, ProductTag.tag_id == Tag.id)
                .where(ProductTag.product_id == product.id)
            )
        ).scalars().all()
        tags = [f"#{tag}" for tag in tag_result] if tag_result else []
    except Exception:
        # Si falla la consulta de tags, continuar sin tags
        tags = []

    # Obtener imgenes del producto (para Telegram/Chatbot)
    images = []
    try:
        img_result = (
            await session.execute(
                select(Image)
                .where(Image.product_id == product.id)
                .where(Image.active == True)
                .order_by(Image.sort_order, Image.id)
            )
        ).scalars().all()
        for img in img_result:
            images.append({
                "image_id": img.id,
                "url": img.url,
                "path": img.path,
                "is_primary": img.is_primary
            })
    except Exception:
        pass

    return {
        "product_id": product.id,
        "sku": canonical_sku,  # SKU canónico (puede ser None si no hay)
        "name": stylize_product_name(canonical_name or product.title) or "(sin nombre)",
        "sale_price": sale_price,
        "stock": stock,
        "description": getattr(product, 'description_html', None),
        "technical_specs": getattr(product, 'technical_specs', None),
        "usage_instructions": getattr(product, 'usage_instructions', None),
        "tags": tags,  # Lista de tags formateados como ["#Organico", "#Floracion"]
        "images": images, # Lista de imágenes
    }


# ------------------------------- Variants Lookup (para MCP Products) -------------------------------
@router.get(
    "/variants/lookup",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def variants_lookup(
    sku: str = Query(None, description="SKU del producto a buscar (canónico preferido)"),
    product_id: int = Query(None, description="ID del producto interno"),
    session: AsyncSession = Depends(get_session),
):
    """Busca un producto por SKU canónico o ID y devuelve información completa.

    Este endpoint es usado por el servidor MCP de productos para obtener
    información detallada de un producto específico.

    Parámetros (usar uno u otro):
      - sku: SKU canónico (formato XXX_####_YYY) o interno
      - product_id: ID interno del producto

    Búsqueda en orden:
      1. Por product_id si se proporciona
      2. Por SKU canónico (CanonicalProduct.sku_custom o ng_sku)
      3. Por SKU interno (Product.sku_root)

    Devuelve: sku (canónico), name, sale_price, stock, description, technical_specs, usage_instructions.
    """
    if not sku and not product_id:
        raise HTTPException(status_code=400, detail={"code": "missing_sku_or_product_id"})

    # 0. Buscar por product_id directamente
    if product_id:
        product = await session.get(Product, product_id)
        if product:
            return await _build_product_response(session, product)
        raise HTTPException(status_code=404, detail={"code": "product_not_found", "product_id": product_id})

    sku_lower = sku.strip().lower()
    if not sku_lower:
        raise HTTPException(status_code=400, detail={"code": "empty_sku"})

    # 1. PRIORIDAD: Buscar por SKU canónico (formato XXX_####_YYY)
    canonical = (
        await session.execute(
            select(CanonicalProduct).where(
                or_(
                    func.lower(CanonicalProduct.sku_custom) == sku_lower,
                    func.lower(CanonicalProduct.ng_sku) == sku_lower,
                )
            )
        )
    ).scalars().first()

    if canonical:
        # Encontrar el producto vinculado al canónico
        product_row = (
            await session.execute(
                select(Product)
                .join(SupplierProduct, SupplierProduct.internal_product_id == Product.id)
                .join(ProductEquivalence, ProductEquivalence.supplier_product_id == SupplierProduct.id)
                .where(ProductEquivalence.canonical_product_id == canonical.id)
                .limit(1)
            )
        ).scalars().first()
        
        if product_row:
            return await _build_product_response(session, product_row)
        
        # Si no hay producto vinculado, devolver info del canónico
        return {
            "product_id": None,
            "sku": canonical.sku_custom or canonical.ng_sku,
            "name": stylize_product_name(canonical.name) or "(sin nombre)",
            "sale_price": float(canonical.sale_price) if canonical.sale_price else None,
            "stock": 0,
            "description": None,
            "technical_specs": None,
            "usage_instructions": None,
            "tags": [],
        }

    # 2. Buscar por SKU canónico en Product (Product.canonical_sku) - preferido
    product = (
        await session.execute(
            select(Product).where(func.lower(Product.canonical_sku) == sku_lower)
        )
    ).scalars().first()
    
    # 3. Si no se encontró, buscar por sku_root como fallback temporal
    if not product:
        product = (
            await session.execute(
                select(Product).where(func.lower(Product.sku_root) == sku_lower)
            )
        ).scalars().first()

    if product:
        return await _build_product_response(session, product)

    # 3. Buscar en Variant por sku - último recurso
    variant = (
        await session.execute(
            select(Variant).where(func.lower(Variant.sku) == sku_lower)
        )
    ).scalars().first()

    if variant and variant.product_id:
        parent_product = await session.get(Product, variant.product_id)
        if parent_product:
            return await _build_product_response(session, parent_product)

    # No encontrado
    raise HTTPException(status_code=404, detail={"code": "product_not_found", "sku": sku})


# ------------------------------- SupplierProduct: link ↔ Variant (upsert) -------------------------------
class _SupplierProductLink(_PydModel):
    supplier_id: int
    supplier_product_id: str
    internal_variant_id: int
    title: Optional[str] = None


@router.post(
    "/supplier-products/link",
    dependencies=[Depends(require_csrf), Depends(require_roles("colaborador", "admin"))],
)
async def supplier_product_link(
    payload: _SupplierProductLink,
    session: AsyncSession = Depends(get_session),
):
    """Crea o actualiza el vínculo entre un SKU de proveedor y una variante interna.

    - Si (supplier_id, supplier_product_id) existe, se actualiza `internal_product_id/internal_variant_id` y `title` opcional.
    - Si no existe, se crea `SupplierProduct` con título opcional.
    - Devuelve el registro resultante con su ID.
    """
    # Validaciones básicas
    supplier = await session.get(Supplier, payload.supplier_id)
    if not supplier:
        raise HTTPException(status_code=400, detail={"code": "invalid_supplier_id"})
    variant = await session.get(Variant, payload.internal_variant_id)
    if not variant:
        raise HTTPException(status_code=400, detail={"code": "invalid_variant_id"})

    product = await session.get(Product, variant.product_id)
    if not product:
        raise HTTPException(status_code=400, detail={"code": "invalid_internal_product"})

    # Buscar existente por clave única (supplier_id, supplier_product_id)
    existing = (
        await session.execute(
            select(SupplierProduct).where(
                (SupplierProduct.supplier_id == payload.supplier_id)
                & (SupplierProduct.supplier_product_id == payload.supplier_product_id)
            )
        )
    ).scalars().first()

    if existing:
        existing.internal_product_id = product.id
        existing.internal_variant_id = variant.id
        if payload.title:
            existing.title = payload.title[:200]
        await session.flush()
        sp = existing
    else:
        sp = SupplierProduct(
            supplier_id=payload.supplier_id,
            supplier_product_id=payload.supplier_product_id,
            title=(payload.title[:200] if payload.title else variant.name or product.title)[:200],
            internal_product_id=product.id,
            internal_variant_id=variant.id,
        )
        session.add(sp)
        await session.flush()

    await session.commit()
    return {
        "id": sp.id,
        "supplier_id": sp.supplier_id,
        "supplier_product_id": sp.supplier_product_id,
        "title": sp.title,
        "internal_product_id": sp.internal_product_id,
        "internal_variant_id": sp.internal_variant_id,
    }


class _ProductsDeleteReq(_PydModel):
    ids: List[int]


@router.delete(
    "/catalog/products",
    dependencies=[Depends(require_csrf), Depends(require_roles("colaborador", "admin"))],
)
async def delete_products_guarded(payload: _ProductsDeleteReq, session: AsyncSession = Depends(get_session)):
    """Elimina productos si no tienen stock ni referencias en compras.

    Respuestas:
    - 400 si alguno tiene stock > 0 (single) con detalle.
    - 409 si está referenciado por líneas de compra.
    - 200 con resumen en otros casos.
    """
    blocked_stock: list[int] = []
    blocked_refs: list[int] = []
    deleted: list[int] = []
    for pid in payload.ids:
        p = await session.get(Product, pid)
        if not p:
            continue
        if int(p.stock or 0) > 0:
            blocked_stock.append(pid)
            continue
        ref = await session.scalar(select(func.count()).select_from(PurchaseLine).where(
            (PurchaseLine.product_id == pid)
        ))
        if (ref or 0) > 0:
            blocked_refs.append(pid)
            continue
        # Eliminar explícitamente dependencias para compatibilidad con motores sin ON DELETE CASCADE
        # 1) SupplierProduct vinculados
        sp_count = 0
        sph_count = 0
        try:
            sps = (await session.execute(select(SupplierProduct).where(SupplierProduct.internal_product_id == pid))).scalars().all()
            for sp in sps:
                # Borrar histories primero (FK sin ON DELETE CASCADE)
                try:
                    sph_list = (await session.execute(select(SupplierPriceHistory).where(SupplierPriceHistory.supplier_product_fk == sp.id))).scalars().all()
                    for sph in sph_list:
                        await session.delete(sph)
                        sph_count += 1
                except Exception:
                    pass
                await session.delete(sp)
                sp_count += 1
        except Exception:
            sp_count = 0
        # 2) Variants e Inventory
        var_count = 0
        inv_count = 0
        try:
            vars = (await session.execute(select(Variant).where(Variant.product_id == pid))).scalars().all()
            for v in vars:
                inv = await session.scalar(select(Inventory).where(Inventory.variant_id == v.id))
                if inv:
                    await session.delete(inv)
                    inv_count += 1
                await session.delete(v)
                var_count += 1
        except Exception:
            pass
        # 3) Imágenes
        img_count = 0
        try:
            imgs = (await session.execute(select(Image).where(Image.product_id == pid))).scalars().all()
            for im in imgs:
                await session.delete(im)
                img_count += 1
        except Exception:
            pass
        # 4) Producto
        await session.delete(p)
        # 5) AuditLog por producto
        try:
            session.add(AuditLog(action="product_delete", table="products", entity_id=pid, meta={
                "cascade": {"supplier_products": sp_count, "supplier_price_history": sph_count, "variants": var_count, "inventories": inv_count, "images": img_count}
            }))
        except Exception:
            pass
        deleted.append(pid)
    await session.commit()
    if len(payload.ids) == 1 and blocked_stock:
        raise HTTPException(status_code=400, detail={"code": "product_has_stock", "message": "Producto con stock no puede eliminarse"})
    if len(payload.ids) == 1 and blocked_refs:
        raise HTTPException(status_code=409, detail={"code": "product_has_references", "message": "Producto referenciado por compras"})
    return {"requested": payload.ids, "deleted": deleted, "blocked_stock": blocked_stock, "blocked_refs": blocked_refs}



# ------------------------------- Proveedores -------------------------------


class SupplierCreate(BaseModel):
    slug: str
    name: str
    location: Optional[str] = None
    contact_name: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    notes: Optional[str] = None
    extra_json: Optional[dict] = None


class SupplierUpdate(BaseModel):
    name: str
    location: Optional[str] = None
    contact_name: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    notes: Optional[str] = None
    extra_json: Optional[dict] = None


ALLOWED_SUPPLIER_FILE_EXT = {"pdf", "txt", "csv", "xls", "xlsx", "ods", "png", "jpg", "jpeg", "webp"}
MAX_SUPPLIER_FILE_BYTES = int(os.getenv("SUPPLIER_FILE_MAX_BYTES", "10485760"))  # 10MB por defecto
SUPPLIER_FILES_ROOT = Path(os.getenv("SUPPLIER_FILES_ROOT", "data/suppliers"))


@router.get(
    "/suppliers",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def list_suppliers(
    session: AsyncSession = Depends(get_session),
) -> List[dict]:
    """Lista proveedores con estadísticas básicas."""

    result = await session.execute(
        select(
            Supplier,
            func.count(SupplierFile.id).label("files_count"),
            func.max(SupplierFile.uploaded_at).label("last_upload"),
        )
        .outerjoin(SupplierFile, SupplierFile.supplier_id == Supplier.id)
        .group_by(Supplier.id)
        # Ordenar por más reciente primero para que el último proveedor creado
        # aparezca al principio (facilita tests que toman el primer ID)
        .order_by(Supplier.id.desc())
    )
    rows = result.all()
    return [
        {
            "id": supplier.id,
            "slug": supplier.slug,
            "name": supplier.name,
            "created_at": supplier.created_at.isoformat(),
            "last_upload_at": last_upload.isoformat() if last_upload else None,
            "files_count": files_count,
        }
        for supplier, files_count, last_upload in rows
    ]


@router.get(
    "/suppliers/{supplier_id}",
    dependencies=[
        Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))
    ],
)
async def get_supplier(
    supplier_id: int, session: AsyncSession = Depends(get_session)
) -> dict:
    supplier = await session.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    return {
        "id": supplier.id,
        "slug": supplier.slug,
        "name": supplier.name,
        "location": supplier.location,
        "contact_name": supplier.contact_name,
        "contact_email": supplier.contact_email,
        "contact_phone": supplier.contact_phone,
        "notes": supplier.notes,
        "extra_json": supplier.extra_json,
        "created_at": supplier.created_at.isoformat(),
    }


@router.get(
    "/suppliers/{supplier_id}/files",
    dependencies=[
        Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))
    ],
)
async def list_supplier_files(
    supplier_id: int, session: AsyncSession = Depends(get_session)
) -> List[dict]:
    """Lista archivos subidos por un proveedor."""

    result = await session.execute(
        select(SupplierFile)
        .where(SupplierFile.supplier_id == supplier_id)
        .order_by(SupplierFile.uploaded_at.desc())
    )
    files = result.scalars().all()
    return [
        {
            "id": f.id,
            "filename": f.filename,
            "original_name": f.original_name or f.filename,
            "sha256": f.sha256,
            "rows": f.rows,
            "processed": f.processed,
            "dry_run": f.dry_run,
            "uploaded_at": f.uploaded_at.isoformat(),
            "content_type": f.content_type,
            "size_bytes": f.size_bytes,
        }
        for f in files
    ]


@router.post(
    "/suppliers/{supplier_id}/files/upload",
    dependencies=[Depends(require_roles("admin", "colaborador")), Depends(require_csrf)],
)
async def upload_supplier_file(
    supplier_id: int,
    file: UploadFile = File(...),
    notes: Optional[str] = Form(None),
    session: AsyncSession = Depends(get_session),
):
    sup = await session.get(Supplier, supplier_id)
    if not sup:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    original_name = file.filename or "archivo"
    ext = original_name.rsplit('.', 1)[-1].lower() if '.' in original_name else ''
    if ext not in ALLOWED_SUPPLIER_FILE_EXT:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido")
    data = await file.read()
    size = len(data)
    if size > MAX_SUPPLIER_FILE_BYTES:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande")
    sha256 = hashlib.sha256(data).hexdigest()
    existing = await session.scalar(
        select(SupplierFile).where(
            SupplierFile.supplier_id == supplier_id,
            SupplierFile.sha256 == sha256,
        )
    )
    if existing:
        return {
            "id": existing.id,
            "filename": existing.filename,
            "original_name": existing.original_name or existing.filename,
            "uploaded_at": existing.uploaded_at.isoformat(),
            "sha256": existing.sha256,
            "size_bytes": existing.size_bytes,
            "content_type": existing.content_type,
            "processed": existing.processed,
            "dry_run": existing.dry_run,
            "rows": existing.rows,
            "duplicate": True,
        }
    SUPPLIER_FILES_ROOT.mkdir(parents=True, exist_ok=True)
    supplier_dir = SUPPLIER_FILES_ROOT / str(supplier_id)
    supplier_dir.mkdir(parents=True, exist_ok=True)
    safe_base = sha256[:12] + ('.' + ext if ext else '')
    disk_name = safe_base
    path = supplier_dir / disk_name
    with open(path, 'wb') as fh:
        fh.write(data)
    sf = SupplierFile(
        supplier_id=supplier_id,
        filename=disk_name,
        original_name=original_name[:255],
        content_type=file.content_type,
        size_bytes=size,
        sha256=sha256,
        rows=0,
        dry_run=True,
        processed=False,
        notes=notes,
    )
    session.add(sf)
    await session.commit()
    await session.refresh(sf)
    return {
        "id": sf.id,
        "filename": sf.filename,
        "original_name": sf.original_name,
        "uploaded_at": sf.uploaded_at.isoformat(),
        "sha256": sf.sha256,
        "size_bytes": sf.size_bytes,
        "content_type": sf.content_type,
        "processed": sf.processed,
        "dry_run": sf.dry_run,
        "rows": sf.rows,
    }


@router.get(
    "/suppliers/files/{file_id}/download",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def download_supplier_file(
    file_id: int, session: AsyncSession = Depends(get_session)
):
    sf = await session.get(SupplierFile, file_id)
    if not sf:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    supplier_dir = SUPPLIER_FILES_ROOT / str(sf.supplier_id)
    path = supplier_dir / sf.filename
    if not path.exists():
        raise HTTPException(status_code=410, detail="Archivo ausente en disco")

    def iterfile():
        with open(path, 'rb') as fh:
            while True:
                chunk = fh.read(8192)
                if not chunk:
                    break
                yield chunk

    return StreamingResponse(
        iterfile(),
        media_type=sf.content_type or 'application/octet-stream',
        headers={
            'Content-Disposition': f'attachment; filename="{sf.original_name or sf.filename}"'
        },
    )


@router.post(
    "/suppliers",
    dependencies=[Depends(require_csrf), Depends(require_roles("admin"))],
)
async def create_supplier(
    request: Request, session: AsyncSession = Depends(get_session)
):
    """Crea un nuevo proveedor validando formato y unicidad de ``slug``."""

    if not (request.headers.get("content-type") or "").lower().startswith("application/json"):
        raise HTTPException(
            status_code=415, detail="Content-Type debe ser application/json"
        )
    try:
        payload = SupplierCreate.model_validate(await request.json())
    except ValidationError:
        return JSONResponse(
            status_code=400,
            content={"code": "invalid_payload", "message": "Faltan campos"},
        )

    existing = await session.scalar(
        select(Supplier).where(Supplier.slug == payload.slug)
    )
    if existing:
        # Idempotencia amistosa en tests/uso repetido: si el nombre coincide, devolver 200 con el existente.
        if (existing.name or "").strip() == payload.name.strip():
            return {
                "id": existing.id,
                "slug": existing.slug,
                "name": existing.name,
                "location": existing.location,
                "contact_name": existing.contact_name,
                "contact_email": existing.contact_email,
                "contact_phone": existing.contact_phone,
                "notes": existing.notes,
                "extra_json": existing.extra_json,
                "created_at": existing.created_at.isoformat(),
            }
        return JSONResponse(
            status_code=409,
            content={"code": "slug_conflict", "message": "Slug ya utilizado"},
        )
    supplier = Supplier(
        slug=payload.slug,
        name=payload.name,
        location=payload.location,
        contact_name=payload.contact_name,
        contact_email=payload.contact_email,
        contact_phone=payload.contact_phone,
        notes=payload.notes,
        extra_json=payload.extra_json,
    )
    session.add(supplier)
    await session.commit()
    await session.refresh(supplier)
    return {
        "id": supplier.id,
        "slug": supplier.slug,
        "name": supplier.name,
        "location": supplier.location,
        "contact_name": supplier.contact_name,
        "contact_email": supplier.contact_email,
        "contact_phone": supplier.contact_phone,
        "notes": supplier.notes,
        "extra_json": supplier.extra_json,
        "created_at": supplier.created_at.isoformat(),
    }


@router.delete(
    "/suppliers",
    dependencies=[Depends(require_csrf), Depends(require_roles("admin"))],
)
async def bulk_delete_suppliers(
    request: Request,
    session: AsyncSession = Depends(get_session),
    sess: SessionData = Depends(current_session),
    force_cascade: bool = False,
):
    """Eliminación bulk de proveedores con validación de integridad referencial.
    
    Parámetros:
    - ids: Array de IDs de proveedores a eliminar
    - force_cascade: Si es true, elimina en cascada import_jobs y equivalencias (solo registros no críticos)
    
    Retorna:
    - requested: IDs solicitados
    - deleted: IDs eliminados exitosamente
    - blocked: Proveedores bloqueados con razones, conteos y detalles de registros bloqueantes
    - not_found: IDs no encontrados
    - cascade_deleted: Registros eliminados en cascada (si force_cascade=true)
    """
    if request.headers.get("content-type") != "application/json":
        raise HTTPException(status_code=415, detail="Content-Type debe ser application/json")
    
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON inválido")
    
    ids = body.get("ids")
    if not ids or not isinstance(ids, list):
        raise HTTPException(status_code=400, detail="ids requerido (array)")
    
    if len(ids) > 500:
        raise HTTPException(status_code=400, detail="máx 500 ids por solicitud")
    
    # Permitir force_cascade desde body también
    force_cascade = body.get("force_cascade", force_cascade)
    
    from db.models import Purchase, PurchaseLine, ImportJob, ProductEquivalence
    
    requested = list(ids)
    deleted = []
    blocked = []
    not_found = []
    cascade_deleted = {
        "import_jobs": [],
        "product_equivalences": []
    }
    
    for sid in requested:
        supplier = await session.get(Supplier, sid)
        if not supplier:
            not_found.append(sid)
            continue
        
        # Verificar referencias bloqueantes
        reasons = []
        counts = {}
        blocking_details = {}
        
        # Contar compras
        purchases_count = await session.scalar(
            select(func.count()).select_from(Purchase).where(Purchase.supplier_id == sid)
        )
        if purchases_count > 0:
            reasons.append("tiene_compras")
            counts["purchases"] = purchases_count
            # Obtener IDs de compras para referencia
            purchase_ids = (await session.execute(
                select(Purchase.id).where(Purchase.supplier_id == sid).limit(10)
            )).scalars().all()
            blocking_details["purchases"] = {
                "count": purchases_count,
                "sample_ids": list(purchase_ids),
                "action": "No se pueden eliminar automáticamente. Revisar módulo de compras."
            }
        
        # Contar archivos
        files_count = await session.scalar(
            select(func.count()).select_from(SupplierFile).where(SupplierFile.supplier_id == sid)
        )
        if files_count > 0:
            reasons.append("tiene_archivos")
            counts["files"] = files_count
            file_ids = (await session.execute(
                select(SupplierFile.id).where(SupplierFile.supplier_id == sid).limit(10)
            )).scalars().all()
            blocking_details["files"] = {
                "count": files_count,
                "sample_ids": list(file_ids),
                "action": "Se eliminarán automáticamente (CASCADE). Este bloqueo es informativo."
            }
        
        # Contar import jobs
        import_jobs_count = await session.scalar(
            select(func.count()).select_from(ImportJob).where(ImportJob.supplier_id == sid)
        )
        if import_jobs_count > 0:
            # Obtener detalles de los jobs
            jobs_info = (await session.execute(
                select(ImportJob.id, ImportJob.status).where(ImportJob.supplier_id == sid)
            )).all()
            
            if force_cascade:
                # Eliminar jobs en cascada
                for job_id, _ in jobs_info:
                    job = await session.get(ImportJob, job_id)
                    if job:
                        await session.delete(job)
                        cascade_deleted["import_jobs"].append(job_id)
            else:
                reasons.append("tiene_import_jobs")
                counts["import_jobs"] = import_jobs_count
                blocking_details["import_jobs"] = {
                    "count": import_jobs_count,
                    "jobs": [{"id": jid, "status": status} for jid, status in jobs_info],
                    "action": "Usar force_cascade=true para eliminar automáticamente, o ejecutar: DELETE FROM import_jobs WHERE supplier_id = {}".format(sid)
                }
        
        # Contar equivalencias
        equivalences_count = await session.scalar(
            select(func.count()).select_from(ProductEquivalence).where(ProductEquivalence.supplier_id == sid)
        )
        if equivalences_count > 0:
            equiv_ids = (await session.execute(
                select(ProductEquivalence.id).where(ProductEquivalence.supplier_id == sid)
            )).scalars().all()
            
            if force_cascade:
                # Eliminar equivalencias en cascada
                for eq_id in equiv_ids:
                    eq = await session.get(ProductEquivalence, eq_id)
                    if eq:
                        await session.delete(eq)
                        cascade_deleted["product_equivalences"].append(eq_id)
            else:
                reasons.append("tiene_equivalencias")
                counts["equivalences"] = equivalences_count
                blocking_details["equivalences"] = {
                    "count": equivalences_count,
                    "sample_ids": list(equiv_ids)[:10],
                    "action": "Usar force_cascade=true para eliminar automáticamente, o ejecutar: DELETE FROM product_equivalences WHERE supplier_id = {}".format(sid)
                }
        
        # Contar líneas de compra a través de supplier_products
        sp_ids = (await session.execute(
            select(SupplierProduct.id).where(SupplierProduct.supplier_id == sid)
        )).scalars().all()
        
        purchase_lines_count = 0
        if sp_ids:
            purchase_lines_count = await session.scalar(
                select(func.count()).select_from(PurchaseLine).where(
                    PurchaseLine.supplier_item_id.in_(sp_ids)
                )
            )
            if purchase_lines_count > 0:
                reasons.append("tiene_lineas_compra")
                counts["purchase_lines"] = purchase_lines_count
                pl_ids = (await session.execute(
                    select(PurchaseLine.id).where(PurchaseLine.supplier_item_id.in_(sp_ids)).limit(10)
                )).scalars().all()
                blocking_details["purchase_lines"] = {
                    "count": purchase_lines_count,
                    "sample_ids": list(pl_ids),
                    "action": "No se pueden eliminar automáticamente. Revisar líneas de compra asociadas."
                }
        
        if reasons:
            blocked.append({
                "id": sid,
                "name": supplier.name,
                "reasons": reasons,
                "counts": counts,
                "details": blocking_details
            })
        else:
            # Eliminar supplier_products asociados (si no tienen referencias)
            for sp_id in sp_ids:
                sp_obj = await session.get(SupplierProduct, sp_id)
                if sp_obj:
                    await session.delete(sp_obj)
            
            # Los SupplierFile tienen CASCADE, se eliminan automáticamente
            await session.delete(supplier)
            deleted.append(sid)
            
            # Audit log
            try:
                session.add(
                    AuditLog(
                        action="delete",
                        table="suppliers",
                        entity_id=sid,
                        meta={"name": supplier.name, "slug": supplier.slug},
                        user_id=sess.user.id if sess and sess.user else None,
                        ip=(request.client.host if request and request.client else None),
                    )
                )
            except Exception:
                pass
    
    await session.commit()
    
    # Audit log resumen
    try:
        session.add(
            AuditLog(
                action="suppliers_delete_bulk",
                table="suppliers",
                entity_id=None,
                meta={
                    "requested": len(requested),
                    "deleted": len(deleted),
                    "blocked": len(blocked),
                    "not_found": len(not_found)
                },
                user_id=sess.user.id if sess and sess.user else None,
                ip=(request.client.host if request and request.client else None),
            )
        )
        await session.commit()
    except Exception:
        pass
    
    return {
        "requested": requested,
        "deleted": deleted,
        "blocked": blocked,
        "not_found": not_found,
        "cascade_deleted": cascade_deleted if force_cascade else None,
        "help": {
            "force_cascade": "Agregar 'force_cascade': true al body para eliminar automáticamente import_jobs y product_equivalences",
            "manual_cleanup": "Para bloqueos críticos (compras, líneas), revisar detalles en 'blocked[].details'"
        }
    }


@router.get(
    "/suppliers/{supplier_id}/items",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def supplier_items_lookup(
    supplier_id: int,
    sku_like: Optional[str] = Query(None, min_length=1),
    q: Optional[str] = Query(None),
    limit: int = Query(10, ge=1, le=50),
    session: AsyncSession = Depends(get_session),
) -> List[dict]:
    """Autocomplete de ítems del proveedor por SKU (supplier_product_id) o título.

    - Retorna hasta `limit` resultados con `id`, `supplier_product_id` y `title`.
    """
    stmt = select(SupplierProduct).where(SupplierProduct.supplier_id == supplier_id)
    if sku_like:
        stmt = stmt.where(SupplierProduct.supplier_product_id.ilike(f"%{sku_like}%"))
    if q:
        stmt = stmt.where(SupplierProduct.title.ilike(f"%{q}%"))
    stmt = stmt.limit(limit)
    rows = (await session.execute(stmt)).scalars().all()
    return [
        {
            "id": r.id,
            "supplier_product_id": r.supplier_product_id,
            "title": r.title,
            "product_id": r.internal_product_id,
        }
        for r in rows
    ]


class SupplierItemCreate(BaseModel):
    """Payload para crear una oferta (SupplierProduct) manualmente.

    - supplier_product_id: SKU o identificador del proveedor (obligatorio)
    - title: título descriptivo
    - product_id: id de producto interno a asociar (opcional)
    - purchase_price / sale_price: precios actuales si se desean registrar
    """

    supplier_product_id: str
    title: str
    product_id: Optional[int] = None
    purchase_price: Optional[float] = None
    sale_price: Optional[float] = None


@router.post(
    "/suppliers/{supplier_id}/items",
    dependencies=[Depends(require_csrf), Depends(require_roles("colaborador", "admin"))],
)
async def create_supplier_item(
    supplier_id: int,
    payload: SupplierItemCreate,
    request: Request,
    session_data: SessionData = Depends(current_session),
    session: AsyncSession = Depends(get_session),
):
    """Crea un SupplierProduct manualmente.

    Reglas:
    - Enforce unicidad (supplier_id, supplier_product_id)
    - Si `product_id` se envía, validar que exista el producto.
    - Registra AuditLog con acción `supplier_item_create`.
    """
    supplier = await session.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    existing = await session.scalar(
        select(SupplierProduct).where(
            SupplierProduct.supplier_id == supplier_id,
            SupplierProduct.supplier_product_id == payload.supplier_product_id,
        )
    )
    if existing:
        return JSONResponse(
            status_code=409,
            content={
                "code": "supplier_item_exists",
                "message": "Ya existe un item con ese identificador para el proveedor",
                "id": existing.id,
            },
        )

    internal_product_id: Optional[int] = None
    if payload.product_id is not None:
        prod = await session.get(Product, payload.product_id)
        if not prod:
            raise HTTPException(status_code=400, detail="product_id inválido")
        internal_product_id = prod.id

    sp = SupplierProduct(
        supplier_id=supplier_id,
        supplier_product_id=payload.supplier_product_id.strip(),
        title=payload.title.strip(),
        current_purchase_price=payload.purchase_price,
        current_sale_price=payload.sale_price,
        internal_product_id=internal_product_id,
    )
    session.add(sp)
    await session.commit()
    await session.refresh(sp)

    try:
        session.add(
            AuditLog(
                action="supplier_item_create",
                table="supplier_products",
                entity_id=sp.id,
                meta={
                    "supplier_product_id": sp.supplier_product_id,
                    "title": sp.title,
                    "product_id": sp.internal_product_id,
                    "purchase_price": sp.current_purchase_price,
                    "sale_price": sp.current_sale_price,
                },
                user_id=session_data.user.id if session_data.user else None,
                ip=(request.client.host if request and request.client else None),
            )
        )
        await session.commit()
    except Exception:
        pass

    return {
        "id": sp.id,
        "supplier_product_id": sp.supplier_product_id,
        "title": sp.title,
        "product_id": sp.internal_product_id,
        "purchase_price": float(sp.current_purchase_price) if sp.current_purchase_price is not None else None,
        "sale_price": float(sp.current_sale_price) if sp.current_sale_price is not None else None,
    }


@router.patch(
    "/suppliers/{supplier_id}",
    dependencies=[Depends(require_csrf), Depends(require_roles("admin"))],
)
async def update_supplier(
    supplier_id: int, req: SupplierUpdate, session: AsyncSession = Depends(get_session)
) -> dict:
    """Actualiza el nombre de un proveedor existente."""

    supplier = await session.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    supplier.name = req.name
    supplier.location = req.location
    supplier.contact_name = req.contact_name
    supplier.contact_email = req.contact_email
    supplier.contact_phone = req.contact_phone
    supplier.notes = req.notes
    supplier.extra_json = req.extra_json
    await session.commit()
    await session.refresh(supplier)
    return {
        "id": supplier.id,
        "slug": supplier.slug,
        "name": supplier.name,
        "location": supplier.location,
        "contact_name": supplier.contact_name,
        "contact_email": supplier.contact_email,
        "contact_phone": supplier.contact_phone,
        "notes": supplier.notes,
        "extra_json": supplier.extra_json,
    }


# ------------------------------- Categorías -------------------------------


class CategoryGenRequest(BaseModel):
    file_id: int
    dry_run: bool = True


def _build_category_path(cat: Category, lookup: dict[int, Category]) -> str:
    parts: List[str] = [cat.name]
    parent_id = cat.parent_id
    while parent_id:
        parent = lookup[parent_id]
        parts.append(parent.name)
        parent_id = parent.parent_id
    return ">".join(reversed(parts))


@router.get(
    "/categories",
    dependencies=[
        Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))
    ],
)
async def list_categories(
    kind: Literal["category", "subcategory"] | None = None,
    session: AsyncSession = Depends(get_session),
) -> List[dict]:
    """Lista categorías con su jerarquía completa."""

    stmt = select(Category)
    if kind:
        stmt = stmt.where(Category.kind == kind)
    result = await session.execute(stmt.order_by(Category.name.asc()))
    cats = result.scalars().all()
    return [
        {
            "id": c.id,
            "name": c.name,
            "parent_id": c.parent_id,
            "kind": c.kind,
            "path": c.name,
        }
        for c in cats
    ]


class CategoryCreate(BaseModel):
    name: str
    parent_id: int | None = None
    kind: Literal["category", "subcategory"] | None = None


@router.post(
    "/categories",
    dependencies=[Depends(require_csrf), Depends(require_roles("colaborador", "admin"))],
)
async def create_category(payload: CategoryCreate, session: AsyncSession = Depends(get_session)) -> dict:
    """Crea una categoría. Unicidad por (name, parent_id).

    Respuesta incluye `id`, `name`, `parent_id` y `path` completo.
    """
    name = " ".join((payload.name or "").strip().split())
    if not name:
        raise HTTPException(status_code=400, detail="name requerido")
    kind = payload.kind or ("subcategory" if payload.parent_id else "category")
    # Verificar padre válido (si viene)
    if payload.parent_id:
        parent = await session.get(Category, payload.parent_id)
        if not parent:
            raise HTTPException(status_code=400, detail="parent_id inválido")
    # Unicidad (name, parent_id)
    exists = await session.scalar(
        select(Category).where(Category.kind == kind, func.lower(Category.name) == name.lower())
    )
    if exists:
        raise HTTPException(status_code=409, detail="La categoría ya existe en ese nivel")
    # Crear
    cat = Category(name=name, parent_id=payload.parent_id, kind=kind)
    session.add(cat)
    await session.commit()
    await session.refresh(cat)
    # Calcular path
    # Cargar todas para construir lookup mínimo (ascendentes)
    # Optimización simple: caminar hacia arriba
    parts: list[str] = [cat.name]
    parent_id = cat.parent_id
    while parent_id:
        p = await session.get(Category, parent_id)
        if not p:
            break
        parts.append(p.name)
        parent_id = p.parent_id
    return {"id": cat.id, "name": cat.name, "parent_id": cat.parent_id, "kind": cat.kind, "path": cat.name}


@router.get(
    "/categories/search",
    dependencies=[
        Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))
    ],
)
async def search_categories(
    q: str,
    kind: Literal["category", "subcategory"] | None = None,
    session: AsyncSession = Depends(get_session),
) -> List[dict]:
    """Busca categorías por nombre o path parcial."""

    stmt = select(Category).where(Category.name.ilike(f"%{q}%"))
    if kind:
        stmt = stmt.where(Category.kind == kind)
    result = await session.execute(stmt.order_by(Category.name.asc()))
    cats = result.scalars().all()
    return [
        {
            "id": c.id,
            "name": c.name,
            "parent_id": c.parent_id,
            "kind": c.kind,
            "path": c.name,
        }
        for c in cats
    ]


@router.post(
    "/categories/generate-from-supplier-file",
    dependencies=[Depends(require_csrf), Depends(require_roles("admin"))],
)
async def generate_categories(
    req: CategoryGenRequest, session: AsyncSession = Depends(get_session)
) -> dict:
    """Genera categorías a partir de un archivo de proveedor."""

    # Obtener productos asociados al archivo
    stmt = (
        select(SupplierProduct)
        .join(SupplierPriceHistory, SupplierPriceHistory.supplier_product_fk == SupplierProduct.id)
        .where(SupplierPriceHistory.file_fk == req.file_id)
    )
    result = await session.execute(stmt)
    products = result.scalars().all()

    paths = set()
    for p in products:
        levels = [
            lvl
            for lvl in [p.category_level_1, p.category_level_2, p.category_level_3]
            if lvl
        ]
        if levels:
            paths.add(">".join(levels))

    # Categorías existentes para comparar
    existing_result = await session.execute(select(Category))
    existing_cats = existing_result.scalars().all()
    lookup = {c.id: c for c in existing_cats}
    existing_paths = {_build_category_path(c, lookup) for c in existing_cats}

    proposed = []
    created: List[str] = []
    skipped: List[str] = []

    for path in sorted(paths):
        if path in existing_paths:
            proposed.append({"path": path, "status": "exists"})
            skipped.append(path)
            continue
        proposed.append({"path": path, "status": "new"})
        if req.dry_run:
            continue
        # Crear jerarquía faltante
        parent_id = None
        for name in path.split(">"):
            kind = "category" if parent_id is None else "subcategory"
            q = select(Category).where(
                func.lower(Category.name) == name.lower(), Category.kind == kind
            )
            cat = await session.scalar(q)
            if not cat:
                cat = Category(name=name, parent_id=parent_id, kind=kind)
                session.add(cat)
                await session.flush()
            parent_id = cat.id
        created.append(path)
        existing_paths.add(path)

    if not req.dry_run:
        await session.commit()

    return {"proposed": proposed, "created": created, "skipped": skipped}


# ------------------------------- Productos -------------------------------


class ProductSortBy(str, Enum):
    updated_at = "updated_at"
    precio_venta = "precio_venta"
    precio_compra = "precio_compra"
    name = "name"
    created_at = "created_at"


class SortOrder(str, Enum):
    asc = "asc"
    desc = "desc"


def _get_image_url_for_browser(img: Image, versions: dict) -> str:
    """Obtiene la URL de imagen compatible con navegadores.
    
    Prioriza versiones derivadas en WebP (preferir 'full', luego 'card', luego 'thumb')
    si están disponibles, ya que son compatibles con todos los navegadores.
    Si no hay versiones derivadas, devuelve el original.
    
    Args:
        img: Objeto Image de la base de datos.
        versions: Dict con versiones derivadas {kind: ImageVersion} para esta imagen.
    
    Returns:
        URL normalizada (con forward slashes) para usar en el navegador.
    """
    # Detectar si el archivo original es HEIC/HEIF (no soportado por navegadores)
    is_heic = False
    if img.mime and img.mime.lower() in ("image/heic", "image/heif"):
        is_heic = True
    elif img.path:
        path_lower = img.path.lower()
        if path_lower.endswith(('.heic', '.heif')):
            is_heic = True
    elif img.url:
        url_lower = img.url.lower()
        if url_lower.endswith(('.heic', '.heif')):
            is_heic = True
    
    # Si es HEIC/HEIF o si hay versiones derivadas disponibles, usar versión derivada
    # (las versiones derivadas siempre son WebP, compatibles con navegadores)
    if is_heic or versions:
        # Preferir: full > card > thumb
        for kind in ["full", "card", "thumb"]:
            if kind in versions:
                v = versions[kind]
                if v.path:
                    # Normalizar separadores para URL
                    path_norm = v.path.replace('\\', '/')
                    return f"/media/{path_norm}"
    
    # Fallback: usar URL original (normalizada)
    if img.url:
        return img.url.replace('\\', '/')
    # Si no hay URL, construir desde path
    if img.path:
        path_norm = img.path.replace('\\', '/')
        return f"/media/{path_norm}"
    return img.url or ""


async def _category_path(session: AsyncSession, category_id: int | None) -> str | None:
    if not category_id:
        return None
    parts: List[str] = []
    current_id = category_id
    while current_id:
        cat = await session.get(Category, current_id)
        if not cat:
            break
        parts.append(cat.name)
        current_id = cat.parent_id
    return ">".join(reversed(parts)) if parts else None


async def _taxonomy_path(
    session: AsyncSession,
    category_id: int | None,
    subcategory_id: int | None,
) -> str | None:
    """Compone las dos taxonomías planas sin depender de parent_id."""
    category = await session.get(Category, category_id) if category_id else None
    subcategory = await session.get(Category, subcategory_id) if subcategory_id else None
    parts = [value.name for value in (category, subcategory) if value]
    return " > ".join(parts) if parts else None


async def _stock_export_records(
    session: AsyncSession,
    *,
    supplier_id: int | None,
    category_id: int | None,
    q: str | None,
    stock: str | None,
    created_since_days: int | None,
    sort_by: str,
    order: str,
    product_type: str | None,
) -> list[dict]:
    """Construye el conjunto canónico compartido por las exportaciones de Stock."""
    try:
        sort_by_enum = ProductSortBy(sort_by)
        order_enum = SortOrder(order)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="ordenamiento inválido") from exc

    sp = SupplierProduct
    p = Product
    s = Supplier
    eq = ProductEquivalence
    cp = CanonicalProduct
    stmt = (
        select(sp, p, s, eq, cp)
        .join(s, sp.supplier_id == s.id)
        .join(p, sp.internal_product_id == p.id)
        .outerjoin(eq, eq.supplier_product_id == sp.id)
        .outerjoin(cp, cp.id == eq.canonical_product_id)
    )
    if supplier_id is not None:
        stmt = stmt.where(sp.supplier_id == supplier_id)
    if category_id is not None:
        stmt = stmt.where(p.category_id == category_id)
    if q:
        stmt = stmt.where(or_(p.title.ilike(f"%{q}%"), sp.title.ilike(f"%{q}%")))
    if stock:
        try:
            operator, raw_value = stock.split(":", 1)
            value = Decimal(raw_value)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="stock inválido (use gt:0 o eq:0)") from exc
        if operator == "gt":
            stmt = stmt.where(p.stock > value)
        elif operator == "eq":
            stmt = stmt.where(p.stock == value)
        else:
            raise HTTPException(status_code=400, detail="stock inválido (op debe ser gt o eq)")
    if created_since_days is not None:
        if created_since_days < 0 or created_since_days > 365:
            raise HTTPException(status_code=400, detail="created_since_days fuera de rango (0-365)")
        from datetime import datetime, timedelta
        stmt = stmt.where(p.created_at >= datetime.utcnow() - timedelta(days=created_since_days))
    if product_type and product_type != "all":
        stmt = stmt.where(
            eq.canonical_product_id.is_not(None)
            if product_type == "canonical"
            else eq.canonical_product_id.is_(None)
        )

    sort_map = {
        ProductSortBy.updated_at: sp.last_seen_at,
        ProductSortBy.precio_venta: sp.current_sale_price,
        ProductSortBy.precio_compra: sp.current_purchase_price,
        ProductSortBy.name: p.title,
        ProductSortBy.created_at: p.created_at,
    }
    sort_column = sort_map[sort_by_enum]
    stmt = stmt.order_by(sort_column.asc() if order_enum == SortOrder.asc else sort_column.desc())
    rows = (await session.execute(stmt)).all()

    product_ids = list({product.id for _, product, *_ in rows})
    skus_by_product: dict[int, str | None] = {}
    if product_ids:
        variant_rows = (
            await session.execute(
                select(Variant.product_id, Variant.sku)
                .where(Variant.product_id.in_(product_ids))
                .order_by(Variant.product_id, Variant.id)
            )
        ).all()
        for product_id, sku in variant_rows:
            skus_by_product.setdefault(product_id, sku)

    records: dict[int, dict] = {}
    for supplier_product, product, _supplier, _equivalence, canonical in rows:
        existing = records.get(product.id)
        canonical_price = float(canonical.sale_price) if canonical and canonical.sale_price is not None else None
        if existing:
            if existing["canonical_sale_price"] is None and canonical_price is not None:
                existing["canonical_sale_price"] = canonical_price
            continue
        records[product.id] = {
            "product_id": product.id,
            "name": canonical.name if canonical and canonical.name else product.title,
            "supplier_sale_price": float(supplier_product.current_sale_price) if supplier_product.current_sale_price is not None else None,
            "canonical_sale_price": canonical_price,
            "category_id": canonical.category_id if canonical and canonical.category_id else product.category_id,
            "subcategory_id": canonical.subcategory_id if canonical and canonical.subcategory_id else product.subcategory_id,
            "sku": (canonical.sku_custom or canonical.ng_sku) if canonical else skus_by_product.get(product.id),
            "stock": float(product.stock or 0),
        }

    exported: list[dict] = []
    for record in records.values():
        exported.append({
            "name": stylize_product_name(record["name"]),
            "sale_price": record["canonical_sale_price"] if record["canonical_sale_price"] is not None else record["supplier_sale_price"],
            "category": await _taxonomy_path(session, record["category_id"], record["subcategory_id"]),
            "sku": record["sku"],
            "stock": record["stock"],
        })
    return exported


@router.get(
    "/products",
    dependencies=[
        Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))
    ],
)
async def list_products(
    supplier_id: Optional[int] = None,
    category_id: Optional[int] = None,
    q: Optional[str] = None,
    stock: Optional[str] = None,
    created_since_days: Optional[int] = None,
    page: int = 1,
    page_size: int = 20,
    sort_by: str = "updated_at",
    order: str = "desc",
    type: Optional[str] = Query(None, pattern="^(all|canonical|supplier)$"),
    *,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Lista productos de proveedores con filtros, orden y paginación."""

    max_page = int(os.getenv("PRODUCTS_PAGE_MAX", "100"))
    if page < 1 or page_size < 1 or page_size > max_page:
        raise HTTPException(status_code=400, detail="paginación inválida")

    try:
        sort_by_enum = ProductSortBy(sort_by)
    except ValueError:
        raise HTTPException(status_code=400, detail="sort_by inválido")

    try:
        order_enum = SortOrder(order)
    except ValueError:
        raise HTTPException(status_code=400, detail="order inválido")

    sp = SupplierProduct
    p = Product
    s = Supplier
    eq = ProductEquivalence
    cp = CanonicalProduct

    stmt = (
        select(sp, p, s, eq, cp)
        .join(s, sp.supplier_id == s.id)
        .join(p, sp.internal_product_id == p.id)
        .outerjoin(eq, eq.supplier_product_id == sp.id)
        .outerjoin(cp, cp.id == eq.canonical_product_id)
    )

    if supplier_id is not None:
        stmt = stmt.where(sp.supplier_id == supplier_id)
    if category_id is not None:
        stmt = stmt.where(p.category_id == category_id)
    if q:
        # Búsqueda por nombre interno, título del proveedor, nombre canónico y SKU
        stmt = stmt.where(
            or_(
                p.title.ilike(f"%{q}%"),
                sp.title.ilike(f"%{q}%"),
                cp.name.ilike(f"%{q}%"),
                p.canonical_sku.ilike(f"%{q}%"),
                p.sku_root.ilike(f"%{q}%"),
                cp.ng_sku.ilike(f"%{q}%"),
                cp.sku_custom.ilike(f"%{q}%"),
            )
        )
    # Stock filter: 'gt:0' or 'eq:0'
    if stock:
        try:
            op, val = stock.split(":", 1)
            val_i = int(val)
        except Exception:
            raise HTTPException(status_code=400, detail="stock inválido (use gt:0 o eq:0)")
        if op == "gt":
            stmt = stmt.where(p.stock > val_i)
        elif op == "eq":
            stmt = stmt.where(p.stock == val_i)
        else:
            raise HTTPException(status_code=400, detail="stock inválido (op debe ser gt o eq)")
    # Filtro de productos creados recientemente
    if created_since_days is not None:
        if created_since_days < 0 or created_since_days > 365:
            raise HTTPException(status_code=400, detail="created_since_days fuera de rango (0-365)")
        from datetime import datetime, timedelta
        cutoff = datetime.utcnow() - timedelta(days=created_since_days)
        stmt = stmt.where(p.created_at >= cutoff)

    # Filtro por tipo (canónicos|proveedor|todos)
    if type and type != "all":
        if type == "canonical":
            stmt = stmt.where(eq.canonical_product_id.is_not(None))
        elif type == "supplier":
            stmt = stmt.where(eq.canonical_product_id.is_(None))

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = await session.scalar(count_stmt) or 0

    sort_map = {
        ProductSortBy.updated_at: sp.last_seen_at,
        ProductSortBy.precio_venta: sp.current_sale_price,
        ProductSortBy.precio_compra: sp.current_purchase_price,
        ProductSortBy.name: p.title,
        ProductSortBy.created_at: p.created_at,
    }
    sort_col = sort_map[sort_by_enum]
    sort_col = sort_col.asc() if order_enum == SortOrder.asc else sort_col.desc()

    stmt = (
        stmt.order_by(sort_col)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    result = await session.execute(stmt)
    rows = result.all()

    # Prefetch primer SKU por producto para evitar N+1
    product_ids = [p_obj.id for _, p_obj, *_ in rows]
    skus_by_product: dict[int, str | None] = {}
    if product_ids:
        vs = (
            await session.execute(
                select(Variant.product_id, Variant.sku)
                .where(Variant.product_id.in_(product_ids))
                .order_by(Variant.product_id.asc(), Variant.id.asc())
            )
        ).all()
        for pid, sku in vs:
            if pid not in skus_by_product:
                skus_by_product[pid] = sku
    
    # Prefetch tags por producto para evitar N+1
    from db.models import Tag, ProductTag
    tags_by_product: dict[int, list[dict]] = {}
    if product_ids:
        tag_rows = (
            await session.execute(
                select(ProductTag.product_id, Tag.id, Tag.name)
                .join(Tag, ProductTag.tag_id == Tag.id)
                .where(ProductTag.product_id.in_(product_ids))
                .order_by(ProductTag.product_id.asc(), Tag.name.asc())
            )
        ).all()
        for pid, tag_id, tag_name in tag_rows:
            if pid not in tags_by_product:
                tags_by_product[pid] = []
            tags_by_product[pid].append({"id": tag_id, "name": tag_name})

    # Prefetch images count and primary image URL per product
    images_by_product: dict[int, dict] = {}
    if product_ids:
        from db.models import Image, ImageVersion
        # Get image counts and first active image per product (prioritize is_primary)
        img_rows = (
            await session.execute(
                select(Image.product_id, Image.id, Image.url, Image.path, Image.is_primary)
                .where(Image.product_id.in_(product_ids), Image.active == True)
                .order_by(Image.product_id.asc(), Image.is_primary.desc().nulls_last(), Image.sort_order.asc().nulls_last(), Image.id.asc())
            )
        ).all()
        
        # Group by product_id
        for pid, img_id, img_url, img_path, is_primary in img_rows:
            if pid not in images_by_product:
                images_by_product[pid] = {"count": 0, "primary_id": img_id, "url": img_url, "path": img_path}
            images_by_product[pid]["count"] += 1
        
        # Try to get WebP versions for primary images (prefer thumb for list view)
        primary_ids = [v["primary_id"] for v in images_by_product.values() if v.get("primary_id")]
        if primary_ids:
            version_rows = (
                await session.execute(
                    select(ImageVersion.image_id, ImageVersion.kind, ImageVersion.path)
                    .where(ImageVersion.image_id.in_(primary_ids), ImageVersion.kind.in_(["thumb", "card", "full"]))
                )
            ).all()
            versions_by_img: dict[int, dict] = {}
            for img_id, kind, path in version_rows:
                if img_id not in versions_by_img:
                    versions_by_img[img_id] = {}
                versions_by_img[img_id][kind] = path
            
            # Update URLs to use derived versions
            for pid, img_data in images_by_product.items():
                primary_id = img_data.get("primary_id")
                if primary_id and primary_id in versions_by_img:
                    for kind in ["thumb", "card", "full"]:
                        if kind in versions_by_img[primary_id] and versions_by_img[primary_id][kind]:
                            path_norm = versions_by_img[primary_id][kind].replace('\\', '/')
                            img_data["url"] = f"/media/{path_norm}"
                            break

    items = []
    for sp_obj, p_obj, s_obj, eq_obj, cp_obj in rows:
        cat_path = await _taxonomy_path(session, p_obj.category_id, p_obj.subcategory_id)
        # Estilizar nombre: Title Case con unidades preservadas
        raw_name = cp_obj.name if (cp_obj and getattr(cp_obj, "name", None)) else p_obj.title
        preferred_name = stylize_product_name(raw_name)
        items.append(
            {
                "product_id": p_obj.id,
                "name": stylize_product_name(p_obj.title),
                "preferred_name": preferred_name,
                "supplier": {
                    "id": s_obj.id,
                    "slug": s_obj.slug,
                    "name": s_obj.name,
                },
                "supplier_item_id": sp_obj.id,
                "precio_compra": float(sp_obj.current_purchase_price)
                if sp_obj.current_purchase_price is not None
                else None,
                "precio_venta": float(sp_obj.current_sale_price)
                if sp_obj.current_sale_price is not None
                else None,
                "compra_minima": float(sp_obj.min_purchase_qty)
                if sp_obj.min_purchase_qty is not None
                else None,
                "category_id": p_obj.category_id,
                "subcategory_id": p_obj.subcategory_id,
                "category_path": cat_path,
                "stock": p_obj.stock,
                "updated_at": sp_obj.last_seen_at.isoformat()
                if sp_obj.last_seen_at
                else None,
                "canonical_product_id": eq_obj.canonical_product_id if eq_obj else None,
                "canonical_sale_price": float(cp_obj.sale_price) if (cp_obj and cp_obj.sale_price is not None) else None,
                "canonical_sku": (cp_obj.sku_custom if (cp_obj and cp_obj.sku_custom) else (cp_obj.ng_sku if cp_obj else None)),
                "canonical_name": stylize_product_name(cp_obj.name) if cp_obj else None,
                "first_variant_sku": skus_by_product.get(p_obj.id),
                # Etapa 1: Datos estructurados de enriquecimiento
                "technical_specs": getattr(p_obj, 'technical_specs', None),
                "usage_instructions": getattr(p_obj, 'usage_instructions', None),
                # Tags del producto
                "tags": tags_by_product.get(p_obj.id, []),
                # Images info
                "image_url": images_by_product.get(p_obj.id, {}).get("url"),
                "images_count": images_by_product.get(p_obj.id, {}).get("count", 0),
                "primary_image_id": images_by_product.get(p_obj.id, {}).get("primary_id"),
            }
        )

    return {
        "page": page,
        "page_size": page_size,
        "total": total,
        "items": items,
    }


@router.get(
    "/stock/export.xlsx",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def export_stock_xlsx(
    supplier_id: Optional[int] = None,
    category_id: Optional[int] = None,
    q: Optional[str] = None,
    stock: Optional[str] = None,
    created_since_days: Optional[int] = None,
    sort_by: str = "updated_at",
    order: str = "desc",
    type: Optional[str] = Query(None, pattern="^(all|canonical|supplier)$"),
    *,
    request: Request,
    session: AsyncSession = Depends(get_session),
):
    """Exporta un XLS de stock con columnas: NOMBRE DE PRODUCTO, PRECIO DE VENTA, CATEGORIA, SKU PROPIO.

    Respeta los mismos filtros que /products. El precio de venta prioriza el canónico si existe;
    de lo contrario usa el precio de venta del proveedor.
    """
    records = await _stock_export_records(
        session,
        supplier_id=supplier_id,
        category_id=category_id,
        q=q,
        stock=stock,
        created_since_days=created_since_days,
        sort_by=sort_by,
        order=order,
        product_type=type,
    )

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append(["NOMBRE DE PRODUCTO", "PRECIO DE VENTA", "CATEGORIA", "SKU PROPIO"])
    # Estilos de encabezado: fondo oscuro, texto claro y negrita, centrado
    header_fill = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_alignment = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Completar filas
    max_name_len = 0
    max_cat_len = 0
    max_sku_len = 0
    for rec in records:
        name = rec["name"]
        cat_path = rec["category"]
        precio = rec["sale_price"]
        sku = rec["sku"]
        ws.append([
            name,
            float(precio) if precio is not None else None,
            cat_path or "",
            sku or "",
        ])
        # Negrita para nombre
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        # Actualizar métricas de ancho sugerido
        name_len = len(str(name or ""))
        cat_len = len(str(cat_path or ""))
        sku_len = len(str(sku or ""))
        max_name_len = max(max_name_len, name_len)
        max_cat_len = max(max_cat_len, cat_len)
        max_sku_len = max(max_sku_len, sku_len)

    # Ancho automático para la primera columna (estimación basada en caracteres)
    try:
        ws.column_dimensions['A'].width = min(max(12, max_name_len + 2), 60)
        ws.column_dimensions['C'].width = min(max(12, max_cat_len + 2), 60)
        ws.column_dimensions['D'].width = min(max(12, max_sku_len + 2), 60)
    except Exception:
        pass

    # Serializar
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Correlation id: usar el que inyecta el middleware si está presente
    cid = None
    try:
        # No hay API pública directa; replicamos regla de middleware
        cid = request.headers.get("x-correlation-id") or request.headers.get("x-request-id")
    except Exception:
        cid = None
    headers = {"Content-Disposition": "attachment; filename=stock.xlsx"}
    if cid:
        headers["X-Correlation-Id"] = cid
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get(
    "/stock/export.csv",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def export_stock_csv(
    supplier_id: Optional[int] = None,
    category_id: Optional[int] = None,
    q: Optional[str] = None,
    stock: Optional[str] = None,
    created_since_days: Optional[int] = None,
    sort_by: str = "updated_at",
    order: str = "desc",
    type: Optional[str] = Query(None, pattern="^(all|canonical|supplier)$"),
    session: AsyncSession = Depends(get_session),
):
    records = await _stock_export_records(
        session,
        supplier_id=supplier_id,
        category_id=category_id,
        q=q,
        stock=stock,
        created_since_days=created_since_days,
        sort_by=sort_by,
        order=order,
        product_type=type,
    )
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(["NOMBRE DE PRODUCTO", "PRECIO DE VENTA", "CATEGORIA", "SKU PROPIO"])
    for record in records:
        writer.writerow([record["name"], record["sale_price"], record["category"] or "", record["sku"] or ""])
    content = output.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="stock.csv"'},
    )


@router.get(
    "/stock/export.pdf",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def export_stock_pdf(
    supplier_id: Optional[int] = None,
    category_id: Optional[int] = None,
    q: Optional[str] = None,
    stock: Optional[str] = None,
    created_since_days: Optional[int] = None,
    sort_by: str = "updated_at",
    order: str = "desc",
    type: Optional[str] = Query(None, pattern="^(all|canonical|supplier)$"),
    session: AsyncSession = Depends(get_session),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    records = await _stock_export_records(
        session,
        supplier_id=supplier_id,
        category_id=category_id,
        q=q,
        stock=stock,
        created_since_days=created_since_days,
        sort_by=sort_by,
        order=order,
        product_type=type,
    )
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    data = [["NOMBRE DE PRODUCTO", "PRECIO DE VENTA", "CATEGORIA", "SKU PROPIO"]]
    for record in records:
        price = "" if record["sale_price"] is None else f"$ {record['sale_price']:.2f}"
        data.append([
            Paragraph(str(record["name"]), styles["BodyText"]),
            price,
            Paragraph(str(record["category"] or ""), styles["BodyText"]),
            str(record["sku"] or ""),
        ])
    table = Table(data, repeatRows=1, colWidths=[100 * mm, 35 * mm, 85 * mm, 45 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BBBBBB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F4F4")]),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    document.build([Paragraph("Stock", styles["Title"]), Spacer(1, 5 * mm), table])
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="stock.pdf"'},
    )


@router.get(
    "/stock/export-tiendanegocio.xlsx",
    dependencies=[Depends(require_roles("colaborador", "admin"))],
)
async def export_stock_tiendanegocio_xlsx(
    supplier_id: Optional[int] = None,
    category_id: Optional[int] = None,
    q: Optional[str] = None,
    stock: Optional[str] = None,
    created_since_days: Optional[int] = None,
    sort_by: str = "updated_at",
    order: str = "desc",
    type: Optional[str] = Query(None, pattern="^(all|canonical|supplier)$"),
    *,
    request: Request,
    session: AsyncSession = Depends(get_session),
):
    """Exporta un XLS con el formato de importación de TiendaNegocio.

    Columnas:
    - SKU (OBLIGATORIO)
    - Nombre del producto
    - Precio
    - Oferta
    - Stock
    - Visibilidad (Visible o Oculto)
    - Descripción
    - Peso en KG
    - Alto en CM
    - Ancho en CM
    - Profundidad en CM
    - Nombre de variante #1
    - Opción de variante #1
    - Nombre de variante #2
    - Opción de variante #2
    - Nombre de variante #3
    - Opción de variante #3
    - Categorías > Subcategorías > … > Subcategorías
    """
    # Reutilizar filtros y ordenamiento como en export_stock_xlsx (sin paginar)
    try:
        sort_by_enum = ProductSortBy(sort_by)
    except ValueError:
        raise HTTPException(status_code=400, detail="sort_by inválido")

    try:
        order_enum = SortOrder(order)
    except ValueError:
        raise HTTPException(status_code=400, detail="order inválido")

    sp = SupplierProduct
    p = Product
    s = Supplier
    eq = ProductEquivalence
    cp = CanonicalProduct

    stmt = (
        select(sp, p, s, eq, cp)
        .join(s, sp.supplier_id == s.id)
        .join(p, sp.internal_product_id == p.id)
        .outerjoin(eq, eq.supplier_product_id == sp.id)
        .outerjoin(cp, cp.id == eq.canonical_product_id)
    )

    if supplier_id is not None:
        stmt = stmt.where(sp.supplier_id == supplier_id)
    if category_id is not None:
        stmt = stmt.where(p.category_id == category_id)
    if q:
        stmt = stmt.where(or_(p.title.ilike(f"%{q}%"), sp.title.ilike(f"%{q}%")))
    if stock:
        try:
            op, val = stock.split(":", 1)
            val_i = int(val)
        except Exception:
            raise HTTPException(status_code=400, detail="stock inválido (use gt:0 o eq:0)")
        if op == "gt":
            stmt = stmt.where(p.stock > val_i)
        elif op == "eq":
            stmt = stmt.where(p.stock == val_i)
        else:
            raise HTTPException(status_code=400, detail="stock inválido (op debe ser gt o eq)")
    if created_since_days is not None:
        if created_since_days < 0 or created_since_days > 365:
            raise HTTPException(status_code=400, detail="created_since_days fuera de rango (0-365)")
        from datetime import datetime, timedelta
        cutoff = datetime.utcnow() - timedelta(days=created_since_days)
        stmt = stmt.where(p.created_at >= cutoff)

    if type and type != "all":
        if type == "canonical":
            stmt = stmt.where(eq.canonical_product_id.is_not(None))
        elif type == "supplier":
            stmt = stmt.where(eq.canonical_product_id.is_(None))

    sort_map = {
        ProductSortBy.updated_at: sp.last_seen_at,
        ProductSortBy.precio_venta: sp.current_sale_price,
        ProductSortBy.precio_compra: sp.current_purchase_price,
        ProductSortBy.name: p.title,
        ProductSortBy.created_at: p.created_at,
    }
    sort_col = sort_map[sort_by_enum]
    sort_col = sort_col.asc() if order_enum == SortOrder.asc else sort_col.desc()
    stmt = stmt.order_by(sort_col)

    result = await session.execute(stmt)
    rows = result.all()

    # Prefetch primer SKU por producto
    product_ids = list({p_obj.id for _, p_obj, *_ in rows})
    skus_by_product: dict[int, str | None] = {}
    if product_ids:
        vs = (
            await session.execute(
                select(Variant.product_id, Variant.sku)
                .where(Variant.product_id.in_(product_ids))
                .order_by(Variant.product_id.asc(), Variant.id.asc())
            )
        ).all()
        for pid, sku in vs:
            if pid not in skus_by_product:
                skus_by_product[pid] = sku

    # Mapear por producto priorizando canónicos
    by_product: dict[int, dict] = {}
    for sp_obj, p_obj, s_obj, eq_obj, cp_obj in rows:
        rec = by_product.get(p_obj.id)
        if rec:
            # Completar precio canónico si aún no
            if rec.get("canonical_sale_price") is None and (cp_obj and cp_obj.sale_price is not None):
                rec["canonical_sale_price"] = float(cp_obj.sale_price)
            continue
        by_product[p_obj.id] = {
            "product_id": p_obj.id,
            "name": p_obj.title,
            "stock": p_obj.stock,
            "description_html": getattr(p_obj, "description_html", None),
            "weight_kg": float(p_obj.weight_kg) if p_obj.weight_kg is not None else None,
            "height_cm": float(p_obj.height_cm) if p_obj.height_cm is not None else None,
            "width_cm": float(p_obj.width_cm) if p_obj.width_cm is not None else None,
            "depth_cm": float(p_obj.depth_cm) if p_obj.depth_cm is not None else None,
            "category_id": p_obj.category_id,
            "subcategory_id": p_obj.subcategory_id,
            "supplier_sale_price": float(sp_obj.current_sale_price) if sp_obj.current_sale_price is not None else None,
            "canonical_sale_price": float(cp_obj.sale_price) if (cp_obj and cp_obj.sale_price is not None) else None,
            "canonical_name": (cp_obj.name if (cp_obj and getattr(cp_obj, "name", None)) else None),
            "canonical_sku": (cp_obj.sku_custom or cp_obj.ng_sku) if cp_obj else None,
            "canonical_category_id": getattr(cp_obj, "category_id", None) if cp_obj else None,
            "canonical_subcategory_id": getattr(cp_obj, "subcategory_id", None) if cp_obj else None,
        }

    # Construir workbook con cabecera TiendaNegocio
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append([
        "SKU (OBLIGATORIO)",
        "Nombre del producto",
        "Precio",
        "Oferta",
        "Stock",
        "Visibilidad (Visible o Oculto)",
        "Descripción",
        "Peso en KG",
        "Alto en CM",
        "Ancho en CM",
        "Profundidad en CM",
        "Nombre de variante #1",
        "Opción de variante #1",
        "Nombre de variante #2",
        "Opción de variante #2",
        "Nombre de variante #3",
        "Opción de variante #3",
        "Categorías > Subcategorías > … > Subcategorías",
    ])
    header_fill = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_alignment = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Filas
    for pid, rec in by_product.items():
        # SKU obligatorio: canónico si existe, si no, primer SKU de variante
        sku = rec.get("canonical_sku") or skus_by_product.get(pid) or ""
        # Nombre preferido: estilizado con Title Case
        name = stylize_product_name(rec.get("canonical_name") or rec.get("name") or "")
        # Precio: priorizar canónico
        precio = rec.get("canonical_sale_price") if rec.get("canonical_sale_price") is not None else rec.get("supplier_sale_price")
        # Stock
        stock_val = rec.get("stock") or 0
        # Visibilidad: Visible por defecto
        vis = "Visible"
        # Descripción
        descripcion = rec.get("description_html") or ""
        # Medidas
        weight_kg = rec.get("weight_kg")
        height_cm = rec.get("height_cm")
        width_cm = rec.get("width_cm")
        depth_cm = rec.get("depth_cm")
        # Categoría jerárquica
        can_subcat_id = rec.get("canonical_subcategory_id")
        can_cat_id = rec.get("canonical_category_id")
        cat_path = await _taxonomy_path(
            session,
            can_cat_id or rec.get("category_id"),
            can_subcat_id or rec.get("subcategory_id"),
        )

        ws.append([
            sku,
            name,
            float(precio) if precio is not None else None,
            "",  # Oferta (vacío)
            int(stock_val),
            vis,
            descripcion,
            weight_kg if weight_kg is not None else None,
            height_cm if height_cm is not None else None,
            width_cm if width_cm is not None else None,
            depth_cm if depth_cm is not None else None,
            "", "",  # Variante #1
            "", "",  # Variante #2
            "", "",  # Variante #3
            cat_path or "",
        ])

    # Serializar
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": "attachment; filename=productos_tiendanegocio.xlsx"}
    try:
        cid = request.headers.get("x-correlation-id") or request.headers.get("x-request-id")
        if cid:
            headers["X-Correlation-Id"] = cid
    except Exception:
        pass
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


class StockUpdate(BaseModel):
    stock: Decimal
    expected_stock: Optional[Decimal] = None

    @field_validator("stock", "expected_stock")
    @classmethod
    def validate_stock_precision(cls, value: Decimal | None) -> Decimal | None:
        if value is None:
            return None
        if value.as_tuple().exponent < -2:
            raise ValueError("stock admite como máximo dos decimales")
        if value < 0 or value > Decimal("1000000000"):
            raise ValueError("stock fuera de rango")
        return value.quantize(Decimal("0.01"))


class ProductCreate(BaseModel):
    title: str
    category_id: Optional[int] = None
    subcategory_id: Optional[int] = None
    tag_names: List[str] = Field(default_factory=list)
    initial_stock: Decimal = Decimal("0")
    status: Optional[str] = None
    # Campos para creación de categoría en línea
    new_category_name: Optional[str] = None
    new_category_parent_id: Optional[int] = None
    # Campos opcionales para autocreación de SupplierProduct desde flujo de compras
    supplier_id: Optional[int] = None
    supplier_sku: Optional[str] = None
    # Permite, si se conoce, enlazar directamente con un producto canónico
    canonical_product_id: Optional[int] = None
    # Diagnóstico: contexto de compra (no es necesario validar existencia de línea aquí)
    purchase_id: Optional[int] = None
    purchase_line_index: Optional[int] = None

    def validate_values(self):
        if self.initial_stock < 0:
            raise ValueError("initial_stock debe ser >= 0")


def _slugify(value: str) -> str:
    value = value.lower().strip()
    repl = []
    for ch in value:
        if ch.isalnum():
            repl.append(ch)
        elif ch in {" ", "-", "_"}:
            repl.append("-")
    slug = "".join(repl)
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug.strip("-")[:180]


def _gen_sku_root(title: str) -> str:
    base = ''.join([c for c in title.upper() if c.isalnum()])[:8]
    if not base:
        base = "PRD"
    return base


@router.post(
    "/products",
    dependencies=[Depends(require_csrf), Depends(require_roles("colaborador", "admin"))],
)
async def create_product(
    payload: ProductCreate,
    session: AsyncSession = Depends(get_session),
    request: Request = None,
    sess: SessionData = Depends(current_session),
):
    try:
        payload.validate_values()
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    final_category_id = payload.category_id
    final_subcategory_id = payload.subcategory_id
    created_category_id = None

    # Lógica de creación de categoría en línea
    if payload.new_category_name:
        cat_name = payload.new_category_name.strip()
        if not cat_name:
            raise HTTPException(status_code=400, detail="El nombre de la nueva categoría no puede estar vacío")

        # Validar que el padre exista, si se proveyó
        if payload.new_category_parent_id:
            parent_cat = await session.get(Category, payload.new_category_parent_id)
            if not parent_cat or parent_cat.kind != "category":
                raise HTTPException(status_code=400, detail="La categoría padre seleccionada no existe")

        # Buscar si ya existe una categoría con el mismo nombre y padre
        legacy_kind = "subcategory" if payload.new_category_parent_id else "category"
        existing_cat = await session.scalar(
            select(Category).where(
                Category.kind == legacy_kind,
                func.lower(Category.name) == cat_name.lower(),
            )
        )

        if existing_cat:
            if legacy_kind == "subcategory":
                final_category_id = payload.new_category_parent_id
                final_subcategory_id = existing_cat.id
            else:
                final_category_id = existing_cat.id
        else:
            # Crear la nueva categoría
            new_cat = Category(name=cat_name, parent_id=payload.new_category_parent_id, kind=legacy_kind)
            session.add(new_cat)
            await session.flush() # Flush para obtener el ID
            if legacy_kind == "subcategory":
                final_category_id = payload.new_category_parent_id
                final_subcategory_id = new_cat.id
            else:
                final_category_id = new_cat.id
            created_category_id = new_cat.id
    
    # Validar categoría si se provee y no se creó una nueva
    if final_category_id is not None and not created_category_id:
        cat = await session.get(Category, final_category_id)
        if not cat or cat.kind != "category":
            raise HTTPException(status_code=400, detail="category_id inválido")

    if final_subcategory_id is not None:
        subcategory = await session.get(Category, final_subcategory_id)
        if not subcategory or subcategory.kind != "subcategory":
            raise HTTPException(status_code=400, detail="subcategory_id inválido")

    sku_root = _gen_sku_root(payload.title)
    slug = _slugify(payload.title)
    # Reglas de stock inicial:
    # - Si la creación se hace en contexto de compra (purchase_id) o se pasa supplier_id+supplier_sku
    #   siempre forzamos stock=0 para evitar doble sumatoria (la confirmación aplicará las cantidades).
    force_zero = bool(payload.purchase_id or (payload.supplier_id and payload.supplier_sku))
    initial_stock = 0 if force_zero else payload.initial_stock
    prod = Product(
        sku_root=sku_root,
        title=payload.title,
        category_id=final_category_id,
        subcategory_id=final_subcategory_id,
        status=payload.status or "active",
        slug=slug,
        stock=initial_stock,
    )
    session.add(prod)
    await session.flush()
    if payload.tag_names:
        from db.models import ProductTag
        from services.routers.tags import get_or_create_tags
        for tag in await get_or_create_tags(session, payload.tag_names):
            session.add(ProductTag(product_id=prod.id, tag_id=tag.id))
    await session.commit()
    await session.refresh(prod)
    supplier_product_id = None
    created_supplier_product_id = None
    created_equivalence_id = None
    canonical_product_id = None
    import logging
    logger = logging.getLogger("app")

    # Intentar autocreación de SupplierProduct si vienen datos
    if payload.supplier_id and payload.supplier_sku:
        from db.models import SupplierProduct, ProductEquivalence, CanonicalProduct, Supplier
        # Validar proveedor
        sup = await session.get(Supplier, payload.supplier_id)
        if sup:
            # ¿Existe ya un SupplierProduct con ese SKU?
            existing_sp = await session.scalar(
                select(SupplierProduct).where(
                    SupplierProduct.supplier_id == payload.supplier_id,
                    SupplierProduct.supplier_product_id == payload.supplier_sku,
                )
            )
            if existing_sp and existing_sp.internal_product_id and existing_sp.internal_product_id != prod.id:
                # Evitar sobreescribir vínculo existente; se registrará en log
                supplier_product_id = existing_sp.id
            elif existing_sp:
                # Completar internal_product_id si faltaba
                if not existing_sp.internal_product_id:
                    existing_sp.internal_product_id = prod.id
                    supplier_product_id = existing_sp.id
                    created_supplier_product_id = existing_sp.id  # Se considera actualización
            else:
                # Crear nuevo SupplierProduct
                sp = SupplierProduct(
                    supplier_id=payload.supplier_id,
                    supplier_product_id=payload.supplier_sku,
                    title=payload.title[:200],
                    internal_product_id=prod.id,
                )
                session.add(sp)
                await session.flush()
                supplier_product_id = sp.id
                created_supplier_product_id = sp.id
            # Crear equivalencia canónica opcional
            if payload.canonical_product_id:
                canonical = await session.get(CanonicalProduct, payload.canonical_product_id)
                if canonical and supplier_product_id:
                    canonical_product_id = canonical.id
                    existing_eq = await session.scalar(
                        select(ProductEquivalence).where(
                            ProductEquivalence.supplier_id == payload.supplier_id,
                            ProductEquivalence.supplier_product_id == supplier_product_id,
                        )
                    )
                    if not existing_eq:
                        eq = ProductEquivalence(
                            supplier_id=payload.supplier_id,
                            supplier_product_id=supplier_product_id,
                            canonical_product_id=canonical.id,
                            confidence=1.0,
                            source="auto_create",
                        )
                        session.add(eq)
                        await session.flush()
                        created_equivalence_id = eq.id
        await session.commit()
        await session.refresh(prod)

    # Si se creó o vinculó un SupplierProduct en contexto de compra, inicializar precios
    # Regla: precio de venta = precio de compra efectivo de la línea
    # Nota: la confirmación de compra actualizará nuevamente precio de compra y stock; esto es inicial.
    initialized_prices: dict | None = None
    if (payload.purchase_id is not None) and (payload.supplier_id and payload.supplier_sku) and supplier_product_id:
        try:
            from decimal import Decimal
            from sqlalchemy import select as _select
            from db.models import PurchaseLine, SupplierProduct as _SP
            # Buscar una línea de la compra que coincida por supplier_sku (la primera sin product_id si existe)
            ln = await session.scalar(
                _select(PurchaseLine)
                .where(
                    PurchaseLine.purchase_id == payload.purchase_id,
                    PurchaseLine.supplier_sku == payload.supplier_sku,
                )
                .order_by(PurchaseLine.id.asc())
            )
            if ln:
                disc = Decimal(str(ln.line_discount or 0)) / Decimal("100")
                unit = Decimal(str(ln.unit_cost or 0))
                eff = unit * (Decimal("1") - disc)
                sp_obj = await session.get(_SP, supplier_product_id)
                if sp_obj:
                    # Inicializar ambos precios actuales a partir del costo efectivo
                    sp_obj.current_purchase_price = eff
                    sp_obj.current_sale_price = eff
                    try:
                        import logging
                        logging.getLogger("growen").info(
                            "product_create_ctx default_sale_applied sp=%s eff=%s purchase_id=%s line_id=%s",
                            sp_obj.id,
                            str(eff),
                            payload.purchase_id,
                            getattr(ln, "id", None),
                        )
                    except Exception:
                        pass
                    initialized_prices = {"purchase_price": float(eff), "sale_price": float(eff)}
                    await session.commit()
        except Exception:
            # No bloquear por errores de inicialización de precio
            pass

    # audit + logging estructurado
    try:
        meta_log = {
            "title": prod.title,
            "category_id": prod.category_id,
            "subcategory_id": prod.subcategory_id,
            "tag_names": payload.tag_names,
            "created_category_id": created_category_id,
            "initial_stock_requested": float(payload.initial_stock),
            "initial_stock_final": float(initial_stock),
            "initial_stock_forced_zero": force_zero,
            "auto_link": bool(payload.supplier_id and payload.supplier_sku),
            "supplier_id": payload.supplier_id,
            "supplier_sku": payload.supplier_sku,
            "supplier_product_id": supplier_product_id,
            "created_supplier_product_id": created_supplier_product_id,
            "created_equivalence_id": created_equivalence_id,
            "canonical_product_id": canonical_product_id,
            "purchase_context": {
                "purchase_id": payload.purchase_id,
                "line_index": payload.purchase_line_index,
            },
        }
        try:
            logger.info(
                "product_create: id=%s supplier_id=%s supplier_sku=%s supplier_product_id=%s created_supplier_product_id=%s canonical_product_id=%s purchase_id=%s line_index=%s forced_zero=%s initial_stock=%s",
                prod.id,
                payload.supplier_id,
                payload.supplier_sku,
                supplier_product_id,
                created_supplier_product_id,
                canonical_product_id,
                payload.purchase_id,
                payload.purchase_line_index,
                force_zero,
                initial_stock,
            )
        except Exception:
            pass
        session.add(
            AuditLog(
                action="product_create",
                table="products",
                entity_id=prod.id,
                meta={**meta_log, **({"initialized_prices": initialized_prices} if initialized_prices else {})},
                user_id=sess.user.id if sess and sess.user else None,
                ip=(request.client.host if request and request.client else None),
            )
        )
        await session.commit()
    except Exception:
        await session.rollback()
        await session.refresh(prod)
    return {
        "id": prod.id,
        "title": prod.title,
        "sku_root": prod.sku_root,
        "slug": prod.slug,
        "stock": prod.stock,
        "category_id": prod.category_id,
        "subcategory_id": prod.subcategory_id,
        "status": prod.status,
        "supplier_product_id": supplier_product_id,
        "canonical_product_id": canonical_product_id,
    }


@router.patch(
    "/products/{product_id}/stock",
    dependencies=[Depends(require_csrf), Depends(require_roles("colaborador", "admin"))],
)
async def update_product_stock(
    product_id: int,
    payload: StockUpdate,
    session: AsyncSession = Depends(get_session),
    request: Request = None,
    sess: SessionData = Depends(current_session),
) -> dict:
    prod = await session.scalar(
        select(Product).where(Product.id == product_id).with_for_update()
    )
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    old = Decimal(str(prod.stock or 0)).quantize(Decimal("0.01"))
    if payload.expected_stock is not None and payload.expected_stock != old:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "El stock cambió desde la última lectura",
                "current_stock": float(old),
            },
        )
    prod.stock = payload.stock
    delta = payload.stock - old
    session.add(StockLedger(
        product_id=product_id,
        source_type="manual_adjustment",
        source_id=product_id,
        delta=delta,
        balance_after=payload.stock,
        meta={"old": float(old), "new": float(payload.stock), "user_id": sess.user.id if sess and sess.user else None},
    ))
    session.add(AuditLog(
        action="product_stock_update",
        table="products",
        entity_id=product_id,
        meta={"old": float(old), "new": float(payload.stock), "delta": float(delta)},
        user_id=sess.user.id if sess and sess.user else None,
        ip=(request.client.host if request and request.client else None),
    ))
    await session.commit()
    return {"product_id": product_id, "stock": float(payload.stock)}


# ------------------------------ Producto por id ------------------------------


@router.get(
    "/products/{product_id}",
    dependencies=[Depends(require_roles("guest", "cliente", "proveedor", "colaborador", "admin"))],
)
async def get_product(product_id: int, session: AsyncSession = Depends(get_session)):
    prod = await session.get(Product, product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    canonical_id = await session.scalar(
        select(ProductEquivalence.canonical_product_id)
        .join(SupplierProduct, SupplierProduct.id == ProductEquivalence.supplier_product_id)
        .where(SupplierProduct.internal_product_id == product_id)
        .order_by(ProductEquivalence.id.asc())
        .limit(1)
    )
    cp = await session.get(CanonicalProduct, canonical_id) if canonical_id else None
    linked_products = [prod]
    if cp:
        linked_products = list(
            (
                await session.scalars(
                    select(Product)
                    .join(SupplierProduct, SupplierProduct.internal_product_id == Product.id)
                    .join(ProductEquivalence, ProductEquivalence.supplier_product_id == SupplierProduct.id)
                    .where(ProductEquivalence.canonical_product_id == cp.id)
                    .distinct()
                    .order_by(Product.id.asc())
                )
            ).all()
        ) or [prod]
    linked_ids = list(dict.fromkeys(item.id for item in linked_products))
    imgs = list(
        (
            await session.scalars(
                select(Image)
                .where(Image.product_id.in_(linked_ids), Image.active == True)
                .order_by(Image.is_primary.desc(), Image.sort_order.asc().nulls_last(), Image.id.asc())
            )
        ).all()
    )
    from db.models import ImageVersion
    versions: dict[int, dict] = {}
    if imgs:
        version_rows = (
            await session.scalars(
                select(ImageVersion).where(
                    ImageVersion.image_id.in_([image.id for image in imgs]),
                    ImageVersion.kind.in_(["full", "card", "thumb"]),
                )
            )
        ).all()
        for version in version_rows:
            versions.setdefault(version.image_id, {})[version.kind] = version
    cat_path = await _taxonomy_path(session, prod.category_id, prod.subcategory_id)
    canonical_sale = float(cp.sale_price) if cp and cp.sale_price is not None else None
    supplier_sale_value = await session.scalar(
        select(SupplierProduct.current_sale_price)
        .where(
            SupplierProduct.internal_product_id.in_(linked_ids),
            SupplierProduct.current_sale_price.is_not(None),
        )
        .order_by(SupplierProduct.last_seen_at.desc().nulls_last())
        .limit(1)
    )
    supplier_sale_price = float(supplier_sale_value) if supplier_sale_value is not None else None
    tag_rows = (
        await session.execute(
            select(Tag.id, Tag.name)
            .join(ProductTag, ProductTag.tag_id == Tag.id)
            .where(ProductTag.product_id.in_(linked_ids))
            .distinct()
            .order_by(Tag.name.asc())
        )
    ).all()
    tags = [{"id": tag_id, "name": tag_name} for tag_id, tag_name in tag_rows]
    supplier_rows = (
        await session.execute(
            select(SupplierProduct, Supplier)
            .join(Supplier, Supplier.id == SupplierProduct.supplier_id)
            .where(SupplierProduct.internal_product_id.in_(linked_ids))
            .order_by(SupplierProduct.internal_product_id.asc(), Supplier.name.asc())
        )
    ).all()
    suppliers_by_product: dict[int, list[dict]] = {item_id: [] for item_id in linked_ids}
    for supplier_product, supplier in supplier_rows:
        suppliers_by_product.setdefault(supplier_product.internal_product_id, []).append(
            {
                "supplier_id": supplier.id,
                "supplier_name": supplier.name,
                "supplier_product_id": supplier_product.supplier_product_id,
                "purchase_price": float(supplier_product.current_purchase_price)
                if supplier_product.current_purchase_price is not None
                else None,
                "sale_price": float(supplier_product.current_sale_price)
                if supplier_product.current_sale_price is not None
                else None,
            }
        )
    latest_job = (
        await session.scalar(
            select(CanonicalEnrichmentJob)
            .where(CanonicalEnrichmentJob.canonical_product_id == cp.id)
            .order_by(CanonicalEnrichmentJob.created_at.desc())
            .limit(1)
        )
        if cp
        else None
    )
    content_source = cp or prod
    stock_total = sum(float(item.stock or 0) for item in linked_products)

    return {
        "id": prod.id,
        "title": stylize_product_name(prod.title),
        "preferred_title": stylize_product_name(cp.name if cp else prod.title),
        "slug": prod.slug,
        "stock": float(prod.stock or 0),
        "stock_total": stock_total,
        "sku_root": prod.sku_root,
        "category_path": cat_path,
        "category_id": prod.category_id,
        "subcategory_id": prod.subcategory_id,
        "description_html": content_source.description_html,
        "enrichment_sources_url": None if cp else getattr(prod, "enrichment_sources_url", None),
        "last_enriched_at": content_source.last_enriched_at.isoformat()
        if getattr(content_source, "last_enriched_at", None)
        else None,
        "enriched_by": getattr(content_source, "enriched_by", None),
        "weight_kg": float(content_source.weight_kg) if content_source.weight_kg is not None else None,
        "height_cm": float(content_source.height_cm) if content_source.height_cm is not None else None,
        "width_cm": float(content_source.width_cm) if content_source.width_cm is not None else None,
        "depth_cm": float(content_source.depth_cm) if content_source.depth_cm is not None else None,
        "technical_specs": content_source.technical_specs or {},
        "usage_instructions": content_source.usage_instructions or {},
        "canonical_product_id": canonical_id,
        "canonical_sale_price": canonical_sale,
        "supplier_sale_price": supplier_sale_price,
        "sale_price": canonical_sale if canonical_sale is not None else supplier_sale_price,
        "canonical_sku": (cp.sku_custom or cp.ng_sku) if cp else None,
        "canonical_name": stylize_product_name(cp.name) if cp else None,
        "canonical_status": "ready" if cp else "canonical_required",
        "content_revision": cp.content_revision if cp else None,
        "enrichment": {
            "job_id": latest_job.id,
            "status": latest_job.status,
            "stage": latest_job.stage,
            "applied_fields": latest_job.applied_fields or [],
            "error": latest_job.error_message,
        } if latest_job else None,
        "images": [
            {
                "id": im.id,
                "product_id": im.product_id,
                "url": _get_image_url_for_browser(im, versions.get(im.id, {})),
                "alt_text": im.alt_text,
                "title_text": im.title_text,
                "is_primary": im.is_primary,
                "locked": im.locked,
                "active": im.active,
            }
            for im in imgs
        ],
        "tags": tags,
        "linked_inventory": [
            {
                "product_id": item.id,
                "original_name": stylize_product_name(item.title),
                "sku_root": item.sku_root,
                "stock": float(item.stock or 0),
                "suppliers": suppliers_by_product.get(item.id, []),
                "product_url": f"/productos/{item.id}",
                "stock_url": f"/stock?product_id={item.id}",
            }
            for item in linked_products
        ],
    }


# ------------------------------ Variantes por producto ------------------------------


@router.get(
    "/products/{product_id}/variants",
    dependencies=[Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))],
)
async def list_product_variants(product_id: int, session: AsyncSession = Depends(get_session)):
    """Devuelve las variantes asociadas a un producto interno.

    Respuesta: lista de objetos con `id`, `sku`, `name`, `value`.
    """
    prod = await session.get(Product, product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    rows = (
        await session.execute(
            select(Variant)
            .where(Variant.product_id == product_id)
            .order_by(Variant.id.asc())
        )
    ).scalars().all()
    return [
        {
            "id": v.id,
            "sku": v.sku,
            "name": v.name,
            "value": v.value,
        }
        for v in rows
    ]


class ProductUpdate(BaseModel):
    category_id: int | None = None
    subcategory_id: int | None = None


class ProductsDeleteRequest(BaseModel):
    ids: List[int]
    hard: bool = False  # futuro: permitir soft-delete si se agrega flag


class EnrichMultipleRequest(BaseModel):
    ids: List[int]
    force: bool | None = False


@router.patch(
    "/products/{product_id}",
    dependencies=[Depends(require_csrf), Depends(require_roles("colaborador", "admin"))],
)
async def patch_product(product_id: int, payload: ProductUpdate, session: AsyncSession = Depends(get_session), request: Request = None, sess: SessionData = Depends(current_session)) -> dict:
    prod = await session.get(Product, product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    data = payload.model_dump(exclude_unset=True)
    old_desc = getattr(prod, "description_html", None)
    old_cat = getattr(prod, "category_id", None)
    old_subcat = getattr(prod, "subcategory_id", None)
    if "category_id" in data:
        # Validar existencia (permitir None para desasociar)
        if data["category_id"] is not None:
            cat = await session.get(Category, int(data["category_id"]))
            if not cat or cat.kind != "category":
                raise HTTPException(status_code=400, detail="category_id inválido")
        prod.category_id = int(data["category_id"]) if data["category_id"] is not None else None
    if "subcategory_id" in data:
        if data["subcategory_id"] is not None:
            subcategory = await session.get(Category, int(data["subcategory_id"]))
            if not subcategory or subcategory.kind != "subcategory":
                raise HTTPException(status_code=400, detail="subcategory_id inválido")
        prod.subcategory_id = int(data["subcategory_id"]) if data["subcategory_id"] is not None else None
    await session.commit()
    # audit description change
    try:
        session.add(
            AuditLog(
                action="product_update",
                table="products",
                entity_id=product_id,
                meta={
                    "fields": list(data.keys()),
                    "desc_len_old": (len(old_desc or "") if old_desc is not None else None),
                    "desc_len_new": (len(prod.description_html or "") if prod.description_html is not None else None),
                    **({"category_old": old_cat, "category_new": prod.category_id} if "category_id" in data else {}),
                    **({"subcategory_old": old_subcat, "subcategory_new": prod.subcategory_id} if "subcategory_id" in data else {}),
                },
                user_id=sess.user.id if sess and sess.user else None,
                ip=(request.client.host if request and request.client else None),
            )
        )
        await session.commit()
    except Exception:
        pass
    return {"status": "ok"}


@router.post(
    "/products/enrich-multiple",
    dependencies=[Depends(require_csrf), Depends(require_roles("admin", "colaborador"))],
)
async def enrich_multiple_products(
    payload: EnrichMultipleRequest,
    session: AsyncSession = Depends(get_session),
    request: Request = None,
    sess: SessionData = Depends(current_session),
) -> dict:
    """Encola/ejecuta enriquecimiento para múltiples productos.

    Reglas:
    - Máximo 20 IDs por solicitud.
    - Se ignoran productos sin título.
    - Si `force` es False, se omiten productos ya enriquecidos (description o fuentes).
    - Reutiliza el flujo de `enrich_product` (ejecución inline para MVP).
    """
    ids = list(dict.fromkeys(payload.ids or []))
    if not ids:
        raise HTTPException(status_code=400, detail="ids requerido")
    if len(ids) > 20:
        raise HTTPException(status_code=400, detail="Máximo 20 productos por lote")
    from services.routers.enrichment import (
        create_enrichment_job,
        dispatch_enrichment_job,
        resolve_canonical_id,
    )

    batch_id = hashlib.sha256(
        f"{sess.user.id if sess and sess.user else None}:{','.join(map(str, ids))}".encode("utf-8")
    ).hexdigest()[:32]
    jobs: list[dict] = []
    skipped_ids: list[int] = []
    seen_canonical: set[int] = set()
    for pid in ids:
        canonical_id = await resolve_canonical_id(session, pid)
        if canonical_id is None:
            skipped_ids.append(pid)
            continue
        if canonical_id in seen_canonical:
            continue
        seen_canonical.add(canonical_id)
        job, created = await create_enrichment_job(
            session,
            canonical_id=canonical_id,
            requested_product_id=pid,
            client_request_id=f"legacy-batch:{batch_id}:{canonical_id}",
            scope="full",
            requested_by_user_id=sess.user.id if sess and sess.user else None,
            batch_id=batch_id,
        )
        if created:
            await dispatch_enrichment_job(job, session)
        jobs.append({"product_id": pid, "canonical_product_id": canonical_id, "job_id": job.id})
    return {
        "status": "queued",
        "batch_id": batch_id,
        "jobs": jobs,
        "skipped": [{"product_id": item, "reason": "canonical_required"} for item in skipped_ids],
    }

@router.get(
    "/debug/enrich/{product_id}",
    dependencies=[Depends(require_roles("admin"))],
)
async def debug_enrich_product(
    product_id: int,
    session: AsyncSession = Depends(get_session),
    request: Request = None,
    sess: SessionData = Depends(current_session),
) -> dict:
    """Endpoint de diagnóstico para el flujo de enriquecimiento.

    No persiste cambios. Devuelve:
    - título elegido (incluye preferencia canónica si aplica)
    - proveedor IA seleccionado y flags relevantes
    - estado de salud del MCP web-search (si está habilitado)
    - prompt generado
    - vista previa de la respuesta de IA (sin parsear)
    """
    prod = await session.get(Product, product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    # Elegir título (preferir canónico): title_canonical -> CanonicalProduct.name por canonical_sku -> product.title
    title = (getattr(prod, "title_canonical", None) or "").strip()
    used_canonical_title = False
    if title:
        used_canonical_title = True
    else:
        try:
            if getattr(prod, "canonical_sku", None):
                cp = (
                    await session.execute(
                        select(CanonicalProduct).where(
                            or_(
                                CanonicalProduct.sku_custom == prod.canonical_sku,
                                CanonicalProduct.ng_sku == prod.canonical_sku,
                            )
                        )
                    )
                ).scalars().first()
                if cp and (cp.name or "").strip():
                    title = cp.name.strip()
                    used_canonical_title = True
        except Exception:
            pass
    if not title:
        title = (prod.title or "").strip()

    # Armar prompt base (mismo que enrich_product)
    schema_hint = (
        "{"
        "\"Título del Producto\": string, "
        "\"Descripción para Nice Grow\": string, "
        "\"Peso KG\": number|null, "
        "\"Alto CM\": number|null, "
        "\"Ancho CM\": number|null, "
        "\"Profundidad CM\": number|null, "
        "\"Fuentes\": object|null  "
        "}"
    )
    prompt = (
        "Eres GrowMaster, un asistente de marketing de productos para jardinería y growshops. "
        "Responde ÚNICAMENTE en JSON válido (sin texto extra, sin markdown, sin ```). "
        "Completa el siguiente esquema con la mejor información posible, usando tono claro y útil, español latino neutro.\n\n"
        f"Producto: {title}\n\n"
        f"Esquema: {schema_hint}\n\n"
        "Reglas: \n"
        "- Si no estás seguro de un valor numérico, usa null.\n"
        "- No inventes datos técnicos; prioriza precisión.\n"
        "- La 'Descripción para Nice Grow' debe ser breve (2-4 oraciones), clara y orientada a clientes.\n"
        "- No generar ni estimar precios: Mercado es la única autoridad monetaria.\n"
        "- Si dispones de fuentes o referencias, incluye un objeto 'Fuentes' con claves descriptivas y valores URL (http/https)."
    )

    # Salud y resultados de MCP Web Search. MCP Products no participa de Enrich.
    web_health = "disabled"
    web_query = None
    web_hits = 0
    web_search_results = None
    try:
        role = getattr(getattr(sess, 'user', None), 'role', 'colaborador') or 'colaborador'
        provider = OpenAIProvider()
        # Web-search si está habilitado
        import os as _os
        use_web = (_os.getenv("AI_USE_WEB_SEARCH", "0").lower() in {"1", "true", "yes"}) and settings.ai_allow_external
        if use_web:
            web_health = "unknown"
            try:
                import httpx as _httpx
                mcp_url = get_mcp_web_search_url()
                health_url = mcp_url.rsplit("/mcp", 1)[0] + "/health"
                async with _httpx.AsyncClient(timeout=2.0) as _cli:
                    _h = await _cli.get(health_url)
                    web_health = "ok" if _h.status_code == 200 else f"bad_status_{_h.status_code}"
            except Exception:
                web_health = "unhealthy"
            if web_health == "ok":
                web_query = title
                try:
                    wres = await provider.call_mcp_web_tool(
                        tool_name="search_web",
                        parameters={
                            "query": web_query,
                            "max_results": int(_os.getenv("AI_WEB_SEARCH_MAX_RESULTS", "3")),
                        },
                        user_role=role,
                    )
                    if isinstance(wres, dict) and wres:
                        items = wres.get("items") or []
                        if isinstance(items, list):
                            web_hits = len(items)
                        web_search_results = wres
                        try:
                            import json as _json
                            prompt += "\n\nBúsqueda web (MCP) - top resultados:\n" + _json.dumps(web_search_results, ensure_ascii=False)
                        except Exception:
                            pass
                except Exception:
                    web_search_results = {"error": "web_search_failed"}
    except Exception:
        pass

    # Provider seleccionado (sin ejecutar cambios)
    router_ai = AIRouter(settings)
    provider_obj = router_ai.get_provider(Task.REASONING.value)
    provider_name = getattr(provider_obj, "name", type(provider_obj).__name__)

    # Ejecutar una llamada de prueba (no persistente) y devolver texto crudo
    try:
        raw = await router_ai.run_async(Task.REASONING.value, prompt)
    except Exception as _e:
        raw = f"<error: {type(_e).__name__}>"

    # Normalización previa de fences y prefijos para test de parseabilidad
    preview = (raw or "")
    norm = preview.strip()
    if norm.startswith("openai:") or norm.startswith("ollama:"):
        norm = norm.split(":", 1)[1].strip()
    if norm.startswith("```"):
        norm = norm.strip("`\n ")
        if norm.lower().startswith("json"):
            norm = norm[4:].strip()
    import json as _json
    will_parse = True
    try:
        _ = _json.loads(norm)
    except Exception:
        will_parse = False

    return {
        "product_id": product_id,
        "title": title,
        "title_used": title,  # alias explícito para claridad de UI/diagnóstico
        "used_canonical_title": used_canonical_title,
        "ai_allow_external": settings.ai_allow_external,
        "ai_provider_selected": provider_name,
        "web_search": {"enabled": bool(web_health != "disabled"), "health": web_health, "query": web_query, "hits": web_hits},
        "prompt": prompt,
        "raw_ai_preview": preview[:1200],
        "raw_ai_looks_json": will_parse,
    }


@router.post(
    "/products/{product_id}/enrich",
    dependencies=[Depends(require_csrf), Depends(require_roles("admin", "colaborador"))],
)
async def enrich_product(
    product_id: int,
    session: AsyncSession = Depends(get_session),
    request: Request = None,
    sess: SessionData = Depends(current_session),
    force: bool = Query(False),
) -> dict:
    """Enriquece un producto usando IA (OpenAI/Ollama vía AIRouter).

    - Requiere rol admin o colaborador.
    - Valida existencia del producto y título.
    - Construye un prompt que solicita JSON con claves conocidas.
    - Actualiza ``description_html`` si viene en la respuesta.
    - Registra ``AuditLog`` con acción ``enrich_ai``.
    """
    prod = await session.get(Product, product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    from services.routers.enrichment import (
        create_enrichment_job,
        dispatch_enrichment_job,
        resolve_canonical_id,
    )

    canonical_product_id = await resolve_canonical_id(session, product_id)
    if canonical_product_id is None:
        raise HTTPException(
            status_code=409,
            detail={"code": "canonical_required", "message": "Debe crear o asignar un canónico"},
        )
    job, created = await create_enrichment_job(
        session,
        canonical_id=canonical_product_id,
        requested_product_id=product_id,
        client_request_id=None if force else f"legacy-product:{product_id}",
        scope="full",
        requested_by_user_id=sess.user.id if sess and sess.user else None,
    )
    if created:
        await dispatch_enrichment_job(job, session)
    return {
        "status": job.status,
        "updated": False,
        "job_id": job.id,
        "canonical_product_id": canonical_product_id,
        "status_url": f"/canonical-products/{canonical_product_id}/enrichment-jobs/{job.id}",
    }

@router.delete(
    "/products/{product_id}/enrichment",
    dependencies=[Depends(require_csrf), Depends(require_roles("admin", "colaborador"))],
)
async def delete_product_enrichment(
    product_id: int,
    session: AsyncSession = Depends(get_session),
    request: Request = None,
    sess: SessionData = Depends(current_session),
) -> dict:
    """Elimina los datos enriquecidos por IA para el producto.

    - Limpia description_html, campos técnicos y enrichment_sources_url.
    - Si existe el archivo .txt de fuentes en MEDIA_ROOT, lo borra.
    - Registra AuditLog con action "delete_enrichment".
    """
    prod = await session.get(Product, product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    from services.routers.enrichment import canonical_snapshot, resolve_canonical_id

    canonical_id = await resolve_canonical_id(session, product_id)
    if canonical_id is None:
        raise HTTPException(
            status_code=409,
            detail={"code": "canonical_required", "message": "El producto no tiene canónico"},
        )
    canonical = await session.get(CanonicalProduct, canonical_id, with_for_update=True)
    if not canonical:
        raise HTTPException(status_code=404, detail="Producto canónico no encontrado")
    previous_snapshot = canonical_snapshot(canonical)
    cleared_fields = [
        field
        for field, value in previous_snapshot.items()
        if value not in (None, {}, [])
    ]
    session.add(
        CanonicalContentVersion(
            canonical_product_id=canonical.id,
            origin="legacy_delete_adapter",
            origin_product_id=product_id,
            revision=canonical.content_revision,
            snapshot_json=previous_snapshot,
            is_applied=False,
            created_by_user_id=sess.user.id if sess and sess.user else None,
        )
    )
    canonical.description_html = None
    canonical.weight_kg = None
    canonical.height_cm = None
    canonical.width_cm = None
    canonical.depth_cm = None
    canonical.technical_specs = {}
    canonical.usage_instructions = {}
    canonical.content_revision += 1
    canonical.last_enriched_at = None
    canonical.enriched_by = None
    session.add(
        AuditLog(
            action="delete_canonical_enrichment",
            table="canonical_products",
            entity_id=canonical.id,
            meta={"requested_product_id": product_id, "fields": cleared_fields},
            user_id=sess.user.id if sess and sess.user else None,
            ip=(request.client.host if request and request.client else None),
        )
    )
    await session.commit()
    return {
        "status": "ok",
        "canonical_product_id": canonical.id,
        "content_revision": canonical.content_revision,
        "cleared_fields": cleared_fields,
        "file_deleted": False,
    }

@router.get(
    "/products/{product_id}/audit-logs",
    dependencies=[Depends(require_roles("colaborador", "admin"))],
)
async def product_audit_logs(product_id: int, session: AsyncSession = Depends(get_session), limit: int = Query(50, ge=1, le=500)) -> dict:
    rows = (await session.execute(
        select(AuditLog)
        .where(AuditLog.table == "products", AuditLog.entity_id == product_id)
        .order_by(AuditLog.created_at.desc())
        .limit(limit)
    )).scalars().all()
    items = [
        {
            "action": r.action,
            "created_at": r.created_at.isoformat() if getattr(r, "created_at", None) else None,
            "meta": r.meta or {},
        }
        for r in rows
    ]
    return {"items": items}


@router.delete(
    "/products",
    dependencies=[Depends(require_csrf), Depends(require_roles("colaborador", "admin"))],
)
async def delete_products(payload: ProductsDeleteRequest, session: AsyncSession = Depends(get_session), request: Request = None, sess: SessionData = Depends(current_session)) -> dict:
    """Borrado sin validaciones estrictas (uso interno/tests).

    A diferencia de ``/catalog/products`` que aplica reglas de stock y referencias,
    este endpoint elimina los productos solicitados de forma directa, junto con
    SupplierProducts asociados, y devuelve la cantidad eliminada.
    """
    if not payload.ids:
        raise HTTPException(status_code=400, detail="ids requerido")
    if len(payload.ids) > 500:
        raise HTTPException(status_code=400, detail="máx 500 ids por solicitud")

    deleted = 0
    for pid in payload.ids:
        prod = await session.get(Product, pid)
        if not prod:
            continue
        # Borrar SupplierProducts asociados (compatibilidad sin ON DELETE CASCADE)
        sp_ids = (await session.execute(select(SupplierProduct.id).where(SupplierProduct.internal_product_id == pid))).scalars().all()
        for sid in sp_ids:
            sp_obj = await session.get(SupplierProduct, sid)
            if sp_obj:
                await session.delete(sp_obj)
        await session.delete(prod)
        deleted += 1
        try:
            session.add(
                AuditLog(
                    action="delete",
                    table="products",
                    entity_id=pid,
                    meta={"name": getattr(prod, "title", None)},
                    user_id=sess.user.id if sess and sess.user else None,
                    ip=(request.client.host if request and request.client else None),
                )
            )
        except Exception:
            pass

    await session.commit()
    # Audit resumen
    try:
        session.add(
            AuditLog(
                action="products_delete_bulk",
                table="products",
                entity_id=None,
                meta={"requested": len(payload.ids), "deleted": deleted},
                user_id=sess.user.id if sess and sess.user else None,
                ip=(request.client.host if request and request.client else None),
            )
        )
        await session.commit()
    except Exception:
        pass
    return {"requested": len(payload.ids), "deleted": deleted}


# --------------------------- Historial de precios --------------------------


class PriceHistoryItem(BaseModel):
    as_of_date: str
    purchase_price: Optional[float]
    sale_price: Optional[float]
    delta_purchase_pct: Optional[float]
    delta_sale_pct: Optional[float]


class PriceHistoryResponse(BaseModel):
    page: int
    page_size: int
    total: int
    items: List[PriceHistoryItem]


@router.get(
    "/price-history",
    dependencies=[
        Depends(require_roles("cliente", "proveedor", "colaborador", "admin"))
    ],
)
async def get_price_history(
    supplier_product_id: Optional[int] = Query(None),
    product_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(DEFAULT_PRICE_HISTORY_PAGE_SIZE, ge=1, le=100),
    session: AsyncSession = Depends(get_session),
) -> PriceHistoryResponse:
    """Devuelve el historial de precios ordenado por fecha."""

    if not supplier_product_id and not product_id:
        raise HTTPException(
            status_code=400,
            detail="Debe indicar supplier_product_id o product_id",
        )

    if supplier_product_id:
        base_query = select(SupplierPriceHistory).where(
            SupplierPriceHistory.supplier_product_fk == supplier_product_id
        )
    else:
        base_query = (
            select(SupplierPriceHistory)
            .join(SupplierProduct)
            .where(SupplierProduct.internal_product_id == product_id)
        )

    total = await session.scalar(
        select(func.count()).select_from(base_query.subquery())
    )

    result = await session.execute(
        base_query.order_by(SupplierPriceHistory.as_of_date.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    rows = result.scalars().all()

    return PriceHistoryResponse(
        page=page,
        page_size=page_size,
        total=total or 0,
        items=[
            PriceHistoryItem(
                as_of_date=r.as_of_date.isoformat(),
                purchase_price=float(r.purchase_price)
                if r.purchase_price is not None
                else None,
                sale_price=float(r.sale_price)
                if r.sale_price is not None
                else None,
                delta_purchase_pct=float(r.delta_purchase_pct)
                if r.delta_purchase_pct is not None
                else None,
                delta_sale_pct=float(r.delta_sale_pct)
                if r.delta_sale_pct is not None
                else None,
            )
            for r in rows
        ],
    )


class PriceUpdate(BaseModel):
    supplier_item_id: int
    purchase_price: Optional[float] = None
    sale_price: Optional[float] = None


@router.patch(
    "/products/{product_id}/prices",
    dependencies=[Depends(require_csrf), Depends(require_roles("colaborador", "admin"))],
)
async def update_product_prices(
    product_id: int,
    payload: PriceUpdate,
    session: AsyncSession = Depends(get_session),
    request: Request = None,
    sess: SessionData = Depends(current_session),
):
    """
    Actualiza precios de un producto.
    - `purchase_price`: Actualiza `current_purchase_price` en `SupplierProduct`.
    - `sale_price`: Actualiza `sale_price` en `CanonicalProduct` si está enlazado.
    """
    prod = await session.get(Product, product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    sp = await session.get(SupplierProduct, payload.supplier_item_id)
    if not sp or sp.internal_product_id != product_id:
        raise HTTPException(status_code=404, detail="Supplier item no encontrado o no corresponde al producto")

    updated_fields = {}
    old_values = {}

    # 1. Actualizar precio de compra del proveedor (SupplierProduct)
    if payload.purchase_price is not None:
        if sp.current_purchase_price != payload.purchase_price:
            old_values["purchase_price"] = sp.current_purchase_price
            sp.current_purchase_price = payload.purchase_price
            updated_fields["purchase_price"] = payload.purchase_price

    # 2. Actualizar precio de venta canónico (CanonicalProduct)
    if payload.sale_price is not None:
        # Encontrar el producto canónico a través de la tabla de equivalencia
        eq = await session.scalar(
            select(ProductEquivalence).where(ProductEquivalence.supplier_product_id == sp.id)
        )
        if eq and eq.canonical_product_id:
            cp = await session.get(CanonicalProduct, eq.canonical_product_id)
            if cp and cp.sale_price != payload.sale_price:
                old_values["sale_price"] = cp.sale_price
                cp.sale_price = payload.sale_price
                updated_fields["sale_price"] = payload.sale_price
        else:
            # Si no hay producto canónico, no se puede actualizar el precio de venta.
            # Podríamos devolver un error o simplemente ignorarlo. Por ahora, lo ignoramos.
            pass

    if not updated_fields:
        return JSONResponse(status_code=304, content={"message": "No changes detected"})

    await session.commit()

    # Registrar en AuditLog
    try:
        session.add(
            AuditLog(
                action="product_price_update",
                table="products",
                entity_id=product_id,
                meta={
                    "supplier_item_id": sp.id,
                    "updated_fields": updated_fields,
                    "old_values": old_values,
                },
                user_id=sess.user.id if sess and sess.user else None,
                ip=(request.client.host if request and request.client else None),
            )
        )
        await session.commit()
    except Exception:
        pass  # No fallar si el loggeo falla

    return {"status": "ok", "updated_fields": updated_fields}
